from flask import Flask, render_template, request, redirect, url_for, flash, jsonify
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
from flask_mail import Mail, Message
from itsdangerous import URLSafeTimedSerializer, SignatureExpired, BadSignature
from datetime import datetime, date, time
import os
import json
import io
import openpyxl
import threading
from collections import defaultdict

app = Flask(__name__)
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'guardias-colegio-secret-2024')
_db_url = os.environ.get('DATABASE_URL', 'sqlite:///guardias.db')
if _db_url.startswith('postgres://'):
    _db_url = _db_url.replace('postgres://', 'postgresql://', 1)
app.config['SQLALCHEMY_DATABASE_URI'] = _db_url
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

# Email config (se configura en settings)
app.config['MAIL_SERVER'] = 'smtp.gmail.com'
app.config['MAIL_PORT'] = 587
app.config['MAIL_USE_TLS'] = True
app.config['MAIL_USERNAME'] = ''
app.config['MAIL_PASSWORD'] = ''
app.config['MAIL_DEFAULT_SENDER'] = ''

db = SQLAlchemy(app)
mail = Mail(app)
login_manager = LoginManager(app)

@app.template_filter('from_json')
def from_json_filter(value):
    try:
        return json.loads(value)
    except Exception:
        return []
login_manager.login_view = 'login'
login_manager.login_message = 'Debes iniciar sesión para acceder.'
serializer = URLSafeTimedSerializer(app.config['SECRET_KEY'])


DIAS_SEMANA = ['Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes']

def parse_etapas(val):
    """Convierte el campo etapa (JSON list o string legacy) a lista."""
    if not val:
        return []
    val = str(val).strip()
    if val.startswith('['):
        try:
            return json.loads(val)
        except Exception:
            return []
    return [val] if val else []
FRANJAS = ['9:00-10:00', '10:00-11:00', 'Patio', '11:30-12:30', '14:30-15:30', '15:30-16:30']

# ───────────────────────────── MODELOS ─────────────────────────────

ETAPAS = ['1-2 años', 'Haur Hezkuntza', 'Lehen Hezkuntza', 'Otras']

class Profesor(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nombre = db.Column(db.String(100), nullable=False)
    email = db.Column(db.String(120), unique=True, nullable=False)
    password_hash = db.Column(db.String(200))
    etapa = db.Column(db.Text, default='[]')  # JSON list: ["Lehen Hezkuntza", "Haur Hezkuntza"]
    aula_tutoria = db.Column(db.String(50))
    aulas_bloqueadas = db.Column(db.Text, default='[]')  # JSON: ["2ºA", "3ºB"]
    es_admin = db.Column(db.Boolean, default=False)
    es_especialista = db.Column(db.Boolean, default=False)
    es_pt = db.Column(db.Boolean, default=False)
    es_educador = db.Column(db.Boolean, default=False)
    activo = db.Column(db.Boolean, default=True)
    de_baja = db.Column(db.Boolean, default=False)
    fecha_baja = db.Column(db.Date)
    fecha_vuelta = db.Column(db.Date)
    horas_max_semanales = db.Column(db.Integer, default=25)
    horas_trabajo_personal = db.Column(db.Integer, default=0)
    horas_libres = db.Column(db.Float, default=0)
    horas_lectivas = db.Column(db.Float, default=0)
    horas_pt = db.Column(db.Float, default=0)
    horas_educador = db.Column(db.Float, default=0)
    materias_especiales = db.Column(db.Text, default='[]')  # JSON: ["Inglés", "Música", ...]
    creado = db.Column(db.DateTime, default=datetime.utcnow)

    horario = db.relationship('HorarioProfesor', backref='profesor', lazy=True, cascade='all, delete-orphan')
    indisponibilidades = db.relationship('Indisponibilidad', backref='profesor', lazy=True, cascade='all, delete-orphan')
    guardias_asignadas = db.relationship('Guardia', foreign_keys='Guardia.profesor_asignado_id', backref='profesor_asignado', lazy=True)
    ausencias = db.relationship('Ausencia', foreign_keys='Ausencia.profesor_id', backref='profesor', lazy=True)

    @property
    def horas_complementarias(self):
        return sum(g.horas_semanales for g in self.grupos)

    @property
    def total_guardias(self):
        return Guardia.query.filter_by(profesor_asignado_id=self.id, completada=True).count()

    def guardias_semana(self, fecha_ref=None):
        if fecha_ref is None:
            fecha_ref = date.today()
        # Lunes y domingo de la semana de fecha_ref
        lunes = fecha_ref - __import__('datetime').timedelta(days=fecha_ref.weekday())
        domingo = lunes + __import__('datetime').timedelta(days=6)
        return Guardia.query.filter(
            Guardia.profesor_asignado_id == self.id,
            Guardia.fecha >= lunes,
            Guardia.fecha <= domingo
        ).count()

    @property
    def porcentaje_guardias(self):
        total_profesores_activos = Profesor.query.filter_by(activo=True, de_baja=False).count()
        if total_profesores_activos == 0:
            return 0
        total_guardias_sistema = Guardia.query.filter_by(completada=True).count()
        if total_guardias_sistema == 0:
            return 0
        esperado = total_guardias_sistema / total_profesores_activos
        if esperado == 0:
            return 0
        return round((self.total_guardias / esperado) * 100, 1)

    def esta_libre_en(self, dia, franja):
        horario = HorarioProfesor.query.filter_by(
            profesor_id=self.id, dia=dia, franja=franja, tiene_clase=True
        ).first()
        if horario:
            return False
        indisponible = Indisponibilidad.query.filter_by(
            profesor_id=self.id, dia=dia, franja=franja, activa=True
        ).first()
        if indisponible:
            return False
        return True


class GrupoTrabajo(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nombre = db.Column(db.String(100), nullable=False, unique=True)
    miembros = db.relationship('ProfesorGrupo', backref='grupo', lazy=True, cascade='all, delete-orphan')


class ProfesorGrupo(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    profesor_id = db.Column(db.Integer, db.ForeignKey('profesor.id'), nullable=False)
    grupo_id = db.Column(db.Integer, db.ForeignKey('grupo_trabajo.id'), nullable=False)
    horas_semanales = db.Column(db.Float, default=1)
    profesor = db.relationship('Profesor', backref=db.backref('grupos', lazy=True))


class SlotComplementaria(db.Model):
    """Franjas de trabajo complementario/personal de cada profesor."""
    id = db.Column(db.Integer, primary_key=True)
    profesor_id = db.Column(db.Integer, db.ForeignKey('profesor.id'), nullable=False)
    dia = db.Column(db.String(20), nullable=False)
    franja = db.Column(db.String(30), nullable=False)
    tipo = db.Column(db.String(20), default='libre')  # 'libre' | 'grupo'
    grupo_id = db.Column(db.Integer, db.ForeignKey('grupo_trabajo.id'), nullable=True)
    tipo2 = db.Column(db.String(20), nullable=True)   # 'libre' when second half is complementaria
    es_manual = db.Column(db.Boolean, default=False)
    profesor = db.relationship('Profesor', backref='slots_complementaria')
    grupo_rel = db.relationship('GrupoTrabajo')


class HorarioProfesor(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    profesor_id = db.Column(db.Integer, db.ForeignKey('profesor.id'), nullable=False)
    dia = db.Column(db.String(20), nullable=False)
    franja = db.Column(db.String(30), nullable=False)
    tiene_clase = db.Column(db.Boolean, default=False)
    asignatura = db.Column(db.String(100))


class Indisponibilidad(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    profesor_id = db.Column(db.Integer, db.ForeignKey('profesor.id'), nullable=False)
    dia = db.Column(db.String(20), nullable=False)
    franja = db.Column(db.String(30), nullable=False)
    motivo = db.Column(db.String(200))
    recurrente = db.Column(db.Boolean, default=False)
    fecha_especifica = db.Column(db.Date)
    activa = db.Column(db.Boolean, default=True)
    creada = db.Column(db.DateTime, default=datetime.utcnow)


class Ausencia(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    profesor_id = db.Column(db.Integer, db.ForeignKey('profesor.id'), nullable=False)
    fecha_inicio = db.Column(db.Date, nullable=False)
    fecha_fin = db.Column(db.Date)
    motivo = db.Column(db.String(200))
    es_baja = db.Column(db.Boolean, default=False)
    notas = db.Column(db.Text)
    reportada_por = db.Column(db.Integer, db.ForeignKey('profesor.id'))
    creada = db.Column(db.DateTime, default=datetime.utcnow)
    # Franjas afectadas (JSON: ["9:00-10:00", "10:00-11:00", ...] o "todas")
    franjas_afectadas = db.Column(db.Text, default='todas')
    # Estado en modo manual: pendiente → el admin aún no ha aprobado
    aprobada = db.Column(db.Boolean, default=True)


class Guardia(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    fecha = db.Column(db.Date, nullable=False)
    dia_semana = db.Column(db.String(20), nullable=False)
    franja = db.Column(db.String(30), nullable=False)
    profesor_ausente_id = db.Column(db.Integer, db.ForeignKey('profesor.id'), nullable=False)
    profesor_asignado_id = db.Column(db.Integer, db.ForeignKey('profesor.id'))
    motivo_ausencia = db.Column(db.String(200))
    estado = db.Column(db.String(20), default='pendiente')  # pendiente, confirmada, rechazada, sin_cobertura
    completada = db.Column(db.Boolean, default=False)
    notas = db.Column(db.Text)
    creada = db.Column(db.DateTime, default=datetime.utcnow)
    aula = db.Column(db.String(50))

    profesor_ausente = db.relationship('Profesor', foreign_keys=[profesor_ausente_id])


class Curso(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nombre = db.Column(db.String(50), unique=True, nullable=False)
    orden = db.Column(db.Integer, default=0)
    etapa = db.Column(db.String(50))
    aula_cerrada = db.Column(db.Boolean, default=False)


class Asignatura(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nombre = db.Column(db.String(100), unique=True, nullable=False)
    color = db.Column(db.String(20), default='#0d6efd')
    etapa = db.Column(db.String(50))


class ProfesorEspecialidad(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    profesor_id = db.Column(db.Integer, db.ForeignKey('profesor.id'), nullable=False)
    nombre = db.Column(db.String(50), nullable=False)
    horas_semanales = db.Column(db.Float, default=0)
    __table_args__ = (db.UniqueConstraint('profesor_id', 'nombre'),)
    profesor = db.relationship('Profesor', backref='especialidades')


class CursoAsignatura(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    curso_id = db.Column(db.Integer, db.ForeignKey('curso.id'), nullable=False)
    asignatura_id = db.Column(db.Integer, db.ForeignKey('asignatura.id'), nullable=False)
    horas_semanales = db.Column(db.Float, default=1)
    __table_args__ = (db.UniqueConstraint('curso_id', 'asignatura_id'),)
    curso = db.relationship('Curso', backref='requisitos')
    asignatura = db.relationship('Asignatura', backref='curso_requisitos')


class ProfesorAsignatura(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    profesor_id = db.Column(db.Integer, db.ForeignKey('profesor.id'), nullable=False)
    asignatura_id = db.Column(db.Integer, db.ForeignKey('asignatura.id'), nullable=False)
    __table_args__ = (db.UniqueConstraint('profesor_id', 'asignatura_id'),)


class HorarioAsignacion(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    curso_id = db.Column(db.Integer, db.ForeignKey('curso.id'), nullable=False)
    asignatura_id = db.Column(db.Integer, db.ForeignKey('asignatura.id'), nullable=False)
    profesor_id = db.Column(db.Integer, db.ForeignKey('profesor.id'), nullable=False)
    dia = db.Column(db.String(20), nullable=False)
    franja = db.Column(db.String(30), nullable=False)
    # Para franjas compartidas (dos medias horas en el mismo slot)
    asignatura2_id = db.Column(db.Integer, db.ForeignKey('asignatura.id'), nullable=True)
    profesor2_id = db.Column(db.Integer, db.ForeignKey('profesor.id'), nullable=True)
    es_manual = db.Column(db.Boolean, default=False)  # True = puesto a mano, no borrar al regenerar
    __table_args__ = (db.UniqueConstraint('curso_id', 'dia', 'franja'),)
    curso = db.relationship('Curso')
    asignatura = db.relationship('Asignatura', foreign_keys=[asignatura_id], overlaps='asignatura2')
    profesor = db.relationship('Profesor', foreign_keys=[profesor_id], overlaps='profesor2')
    asignatura2 = db.relationship('Asignatura', foreign_keys=[asignatura2_id], overlaps='asignatura')
    profesor2 = db.relationship('Profesor', foreign_keys=[profesor2_id], overlaps='profesor')


class ReglaHorario(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    # Reglas de asignatura:
    #   'max_dia', 'consecutivas', 'fijar_franja', 'tutor_primera'
    # Reglas de profesor (dureza 'dura' o 'blanda'):
    #   'prof_excluir_curso', 'prof_fijar_curso',
    #   'prof_excluir_franja', 'prof_fijar_asignatura',
    #   'prof_evitar_curso',  'prof_preferir_curso',
    #   'prof_evitar_franja', 'prof_preferir_asignatura',
    #   'prof_horas_guardia'  (horas de guardia mínimas/máximas)
    tipo = db.Column(db.String(30), nullable=False)
    dureza = db.Column(db.String(10), default='dura')  # 'dura' | 'blanda'
    asignatura_id = db.Column(db.Integer, db.ForeignKey('asignatura.id'), nullable=True)
    profesor_id = db.Column(db.Integer, db.ForeignKey('profesor.id'), nullable=True)
    curso_id_regla = db.Column(db.Integer, db.ForeignKey('curso.id'), nullable=True)
    valor = db.Column(db.Integer, default=1)
    dia = db.Column(db.String(20))
    franja = db.Column(db.String(30))
    etapa = db.Column(db.String(50))
    asignatura = db.relationship('Asignatura', backref='reglas')
    profesor_regla = db.relationship('Profesor', backref='reglas_horario', foreign_keys=[profesor_id])
    curso_regla = db.relationship('Curso', backref='reglas_horario', foreign_keys=[curso_id_regla])


class ConfiguracionEmail(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    mail_server = db.Column(db.String(100), default='smtp.gmail.com')
    mail_port = db.Column(db.Integer, default=587)
    mail_username = db.Column(db.String(120))
    mail_password = db.Column(db.String(200))
    activo = db.Column(db.Boolean, default=False)
    # Modo de gestión de guardias:
    # 'automatico' → el profesor reporta ausencia y se asigna guardia al momento
    # 'manual'     → el admin revisa la ausencia y aprueba la asignación
    modo_guardias = db.Column(db.String(20), default='manual')
    anthropic_api_key = db.Column(db.String(200))


# ───────────────────────────── AUTH ─────────────────────────────

@login_manager.user_loader
def load_user(user_id):
    return Profesor.query.get(int(user_id))


def check_password(stored, provided):
    import hashlib
    return hashlib.sha256(provided.encode()).hexdigest() == stored


def hash_password(password):
    import hashlib
    return hashlib.sha256(password.encode()).hexdigest()


# ───────────────────────────── ALGORITMO ─────────────────────────────

def puntuacion_candidato(profesor, aula_guardia, etapa_ausente, fecha):
    """
    Menor puntuación = mayor prioridad.
    Factores (de mayor a menor peso):
      1. Tutor de la clase donde se hace la guardia  → -100
      2. Misma etapa que el profesor ausente          → -25
      3. Guardias realizadas esta semana              → +15 por guardia
      4. Total de guardias acumuladas                 → +1 por guardia
    """
    score = 0

    # Factor 1: tutor de la clase concreta
    if aula_guardia and profesor.aula_tutoria and \
            profesor.aula_tutoria.strip().lower() == aula_guardia.strip().lower():
        score -= 100

    # Factor 2: misma etapa educativa
    if etapa_ausente and profesor.etapa and profesor.etapa == etapa_ausente:
        score -= 25

    # Factor 3: carga semanal (evitar saturar a alguien esta semana)
    score += profesor.guardias_semana(fecha) * 15

    # Factor 4: equidad acumulada a largo plazo
    score += profesor.total_guardias * 1

    return score


def buscar_profesor_para_guardia(dia, franja, excluir_id, aula=None, etapa_ausente=None, fecha=None):
    if fecha is None:
        fecha = date.today()

    def puede_cubrir(p):
        if p.id == excluir_id or not p.esta_libre_en(dia, franja):
            return False
        if aula:
            bloqueadas = json.loads(p.aulas_bloqueadas or '[]')
            if aula.strip().lower() in [b.strip().lower() for b in bloqueadas]:
                return False
        return True

    candidatos = Profesor.query.filter_by(activo=True, de_baja=False, es_especialista=False, es_pt=False).all()
    candidatos = [p for p in candidatos if puede_cubrir(p)]

    if not candidatos:
        candidatos = Profesor.query.filter_by(activo=True, de_baja=False, es_especialista=True, es_pt=False).all()
        candidatos = [p for p in candidatos if puede_cubrir(p)]

    if not candidatos:
        return None

    candidatos.sort(key=lambda p: puntuacion_candidato(p, aula, etapa_ausente, fecha))
    return candidatos[0]


def get_config():
    cfg = ConfiguracionEmail.query.first()
    if not cfg:
        cfg = ConfiguracionEmail()
        db.session.add(cfg)
        db.session.commit()
    return cfg

def get_mail_config():
    return get_config()


def mail_configurado():
    cfg = get_mail_config()
    return cfg and cfg.activo and cfg.mail_username


def init_mail(cfg):
    app.config['MAIL_USERNAME'] = cfg.mail_username
    app.config['MAIL_PASSWORD'] = cfg.mail_password
    app.config['MAIL_DEFAULT_SENDER'] = cfg.mail_username
    mail.init_app(app)


def enviar_invitacion(profesor):
    cfg = get_mail_config()
    if not cfg or not cfg.activo or not cfg.mail_username:
        return False
    try:
        init_mail(cfg)
        token = serializer.dumps(profesor.email, salt='invitacion-registro')
        url = url_for('registro_invitacion', token=token, _external=True)
        msg = Message(
            subject='Invitación al sistema de guardias — Colegio La Asunción',
            recipients=[profesor.email]
        )
        msg.html = render_template('email_invitacion.html', profesor=profesor, url=url)
        mail.send(msg)
        return True
    except Exception as e:
        print(f'Error enviando invitación: {e}')
        return False


def enviar_email_guardia(guardia):
    cfg = ConfiguracionEmail.query.first()
    if not cfg or not cfg.activo or not cfg.mail_username:
        return False
    try:
        app.config['MAIL_USERNAME'] = cfg.mail_username
        app.config['MAIL_PASSWORD'] = cfg.mail_password
        app.config['MAIL_DEFAULT_SENDER'] = cfg.mail_username
        mail.init_app(app)

        profesor = guardia.profesor_asignado
        ausente = guardia.profesor_ausente
        token_confirmar = serializer.dumps(f'confirmar:{guardia.id}', salt='guardia-action')
        token_rechazar = serializer.dumps(f'rechazar:{guardia.id}', salt='guardia-action')

        url_confirmar = url_for('accion_guardia', token=token_confirmar, _external=True)
        url_rechazar = url_for('accion_guardia', token=token_rechazar, _external=True)

        msg = Message(
            subject=f'Guardia asignada — {guardia.dia_semana} {guardia.franja}',
            recipients=[profesor.email]
        )
        msg.html = render_template('email_guardia.html',
            profesor=profesor, ausente=ausente, guardia=guardia,
            url_confirmar=url_confirmar, url_rechazar=url_rechazar)
        mail.send(msg)
        return True
    except Exception as e:
        print(f'Error email: {e}')
        return False


# ───────────────────────────── RUTAS AUTH ─────────────────────────────


@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        email = request.form.get('email', '').strip().lower()
        password = request.form.get('password', '')
        profesor = Profesor.query.filter_by(email=email).first()
        if profesor and profesor.password_hash and check_password(profesor.password_hash, password):
            login_user(profesor)
            return redirect(url_for('dashboard'))
        flash('Email o contraseña incorrectos.', 'danger')
    return render_template('login.html')


@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))


# ───────────────────────────── DASHBOARD ─────────────────────────────

@app.route('/')
@login_required
def dashboard():
    hoy = date.today()
    guardias_hoy = Guardia.query.filter_by(fecha=hoy).all()
    mis_guardias_pendientes = Guardia.query.filter_by(
        profesor_asignado_id=current_user.id, completada=False
    ).order_by(Guardia.fecha).limit(5).all()
    total_profesores = Profesor.query.filter_by(activo=True, de_baja=False).count()
    de_baja = Profesor.query.filter_by(de_baja=True).count()
    guardias_mes = Guardia.query.filter(
        Guardia.fecha >= date(hoy.year, hoy.month, 1)
    ).count()
    return render_template('dashboard.html',
        guardias_hoy=guardias_hoy,
        mis_guardias=mis_guardias_pendientes,
        total_profesores=total_profesores,
        de_baja=de_baja,
        guardias_mes=guardias_mes,
        hoy=hoy)


# ───────────────────────────── PROFESORES ─────────────────────────────

def _guardar_grupos_profesor(profesor_id, form):
    ProfesorGrupo.query.filter_by(profesor_id=profesor_id).delete()
    for g in GrupoTrabajo.query.all():
        if form.get(f'grupo_{g.id}'):
            horas = float(form.get(f'grupo_horas_{g.id}') or 1)
            db.session.add(ProfesorGrupo(profesor_id=profesor_id, grupo_id=g.id, horas_semanales=horas))


LH_ESPECIFICAS = {'Inglés', 'Música', 'Francés', 'Gimnasia', 'Religión'}
HH_BASE = ['CEA', 'CRR', 'DEE']


def _auto_asignar_asignaturas(profesor):
    """Asigna automáticamente asignaturas según etapa: todas las de primaria salvo las específicas."""
    etapas = parse_etapas(profesor.etapa)

    asig_ids = set()
    if 'Lehen Hezkuntza' in etapas:
        # Todas las asignaturas de primaria excepto las específicas
        for asig in Asignatura.query.filter_by(etapa='Lehen Hezkuntza').all():
            if asig.nombre not in LH_ESPECIFICAS:
                asig_ids.add(asig.id)
    if 'Haur Hezkuntza' in etapas:
        for nombre in HH_BASE:
            asig = Asignatura.query.filter_by(nombre=nombre).first()
            if asig:
                asig_ids.add(asig.id)

    ProfesorAsignatura.query.filter_by(profesor_id=profesor.id).delete()
    for asig_id in asig_ids:
        db.session.add(ProfesorAsignatura(profesor_id=profesor.id, asignatura_id=asig_id))


@app.route('/profesores')
@login_required
def profesores():
    lista = Profesor.query.order_by(Profesor.nombre).all()
    return render_template('profesores.html', profesores=lista)


@app.route('/profesores/nuevo', methods=['GET', 'POST'])
@login_required
def nuevo_profesor():
    if not current_user.es_admin:
        flash('Solo los administradores pueden añadir profesores.', 'danger')
        return redirect(url_for('profesores'))
    if request.method == 'POST':
        email = request.form.get('email', '').strip().lower()
        if not email:
            import uuid
            email = f'sin-email-{uuid.uuid4().hex[:8]}@pendiente.local'
        elif Profesor.query.filter_by(email=email).first():
            flash('Ya existe un profesor con ese email.', 'danger')
            cursos = Curso.query.order_by(Curso.orden, Curso.nombre).all()
            return render_template('form_profesor.html', profesor=None, cursos=cursos)
        aulas_bloqueadas = json.dumps(request.form.getlist('aulas_bloqueadas'))
        p = Profesor(
            nombre=request.form.get('nombre', '').strip(),
            email=email,
            etapa=json.dumps(request.form.getlist('etapas')),
            aula_tutoria=request.form.get('aula_tutoria', '').strip() or None,
            aulas_bloqueadas=aulas_bloqueadas,
            es_admin=bool(request.form.get('es_admin')),
            es_especialista=bool(request.form.get('es_especialista')),
            es_pt=bool(request.form.get('es_pt')),
            es_educador=bool(request.form.get('es_educador')),
            horas_trabajo_personal=int(request.form.get('horas_trabajo_personal', 0) or 0),
            horas_libres=float(request.form.get('horas_libres', 0) or 0),
            horas_lectivas=float(request.form.get('horas_lectivas', 0) or 0),
            horas_pt=float(request.form.get('horas_pt', 0) or 0),
            horas_educador=float(request.form.get('horas_educador', 0) or 0),
            materias_especiales=json.dumps(request.form.getlist('materias_especiales')),
        )
        db.session.add(p)
        db.session.flush()
        _guardar_grupos_profesor(p.id, request.form)
        _auto_asignar_asignaturas(p)
        db.session.commit()
        if mail_configurado():
            ok = enviar_invitacion(p)
            if ok:
                flash(f'Profesor {p.nombre} añadido. Se le ha enviado un email para que cree su contraseña.', 'success')
            else:
                flash(f'Profesor {p.nombre} añadido, pero el email no pudo enviarse. Configura primero el email o usa "Reenviar invitación".', 'warning')
        else:
            flash(f'Profesor {p.nombre} añadido. Configura el email para enviarle la invitación, o usa "Reenviar invitación" más adelante.', 'warning')
        return redirect(url_for('profesores'))
    cursos = Curso.query.order_by(Curso.orden, Curso.nombre).all()
    grupos = GrupoTrabajo.query.order_by(GrupoTrabajo.nombre).all()
    return render_template('form_profesor.html', profesor=None, cursos=cursos, grupos=grupos, grupos_prof={})


@app.route('/profesores/<int:id>/enlace-invitacion')
@login_required
def enlace_invitacion(id):
    if not current_user.es_admin:
        flash('Solo administradores.', 'danger')
        return redirect(url_for('profesores'))
    p = Profesor.query.get_or_404(id)
    token = serializer.dumps(p.email, salt='invitacion-registro')
    enlace = url_for('registro_invitacion', token=token, _external=True)
    return render_template('enlace_invitacion.html', profesor=p, enlace=enlace)


@app.route('/profesores/<int:id>/reenviar-invitacion', methods=['POST'])
@login_required
def reenviar_invitacion(id):
    if not current_user.es_admin:
        flash('Solo administradores.', 'danger')
        return redirect(url_for('profesores'))
    p = Profesor.query.get_or_404(id)
    if p.password_hash:
        flash(f'{p.nombre} ya tiene contraseña establecida.', 'info')
        return redirect(url_for('profesores'))
    ok = enviar_invitacion(p)
    if ok:
        flash(f'Invitación reenviada a {p.email}.', 'success')
    else:
        flash('No se pudo enviar. Comprueba la configuración de email.', 'danger')
    return redirect(url_for('profesores'))


@app.route('/registro/<token>', methods=['GET', 'POST'])
def registro_invitacion(token):
    try:
        email = serializer.loads(token, salt='invitacion-registro', max_age=172800)  # 48h
    except SignatureExpired:
        flash('El enlace ha caducado. Pide al administrador que te reenvíe la invitación.', 'danger')
        return redirect(url_for('login'))
    except BadSignature:
        flash('Enlace no válido.', 'danger')
        return redirect(url_for('login'))

    profesor = Profesor.query.filter_by(email=email).first_or_404()
    if request.method == 'POST':
        password = request.form.get('password', '')
        confirm = request.form.get('confirm', '')
        if len(password) < 6:
            flash('La contraseña debe tener al menos 6 caracteres.', 'danger')
            return render_template('registro_invitacion.html', profesor=profesor, token=token)
        if password != confirm:
            flash('Las contraseñas no coinciden.', 'danger')
            return render_template('registro_invitacion.html', profesor=profesor, token=token)
        profesor.password_hash = hash_password(password)
        db.session.commit()
        login_user(profesor)
        flash(f'¡Bienvenido/a, {profesor.nombre.split()[0]}! Ya puedes usar la aplicación.', 'success')
        return redirect(url_for('dashboard'))
    return render_template('registro_invitacion.html', profesor=profesor, token=token)


@app.route('/profesores/<int:id>/editar', methods=['GET', 'POST'])
@login_required
def editar_profesor(id):
    p = Profesor.query.get_or_404(id)
    if not current_user.es_admin and current_user.id != id:
        flash('No tienes permiso para editar este perfil.', 'danger')
        return redirect(url_for('profesores'))
    if request.method == 'POST':
        p.nombre = request.form.get('nombre', '').strip()
        p.etapa = json.dumps(request.form.getlist('etapas'))
        p.aula_tutoria = request.form.get('aula_tutoria', '').strip() or None
        p.aulas_bloqueadas = json.dumps(request.form.getlist('aulas_bloqueadas'))
        p.horas_libres = float(request.form.get('horas_libres', 0) or 0)
        p.horas_lectivas = float(request.form.get('horas_lectivas', 0) or 0)
        p.horas_pt = float(request.form.get('horas_pt', 0) or 0)
        p.horas_educador = float(request.form.get('horas_educador', 0) or 0)
        p.materias_especiales = json.dumps(request.form.getlist('materias_especiales'))
        if current_user.es_admin:
            p.es_admin = bool(request.form.get('es_admin'))
            p.es_especialista = bool(request.form.get('es_especialista'))
            p.es_pt = bool(request.form.get('es_pt'))
            p.es_educador = bool(request.form.get('es_educador'))
        nueva_pass = request.form.get('password', '').strip()
        if nueva_pass:
            p.password_hash = hash_password(nueva_pass)
        _guardar_grupos_profesor(p.id, request.form)
        _auto_asignar_asignaturas(p)
        db.session.commit()
        flash('Perfil actualizado.', 'success')
        return redirect(url_for('profesores'))
    cursos = Curso.query.order_by(Curso.orden, Curso.nombre).all()
    grupos = GrupoTrabajo.query.order_by(GrupoTrabajo.nombre).all()
    grupos_prof = {m.grupo_id: m.horas_semanales for m in p.grupos}
    return render_template('form_profesor.html', profesor=p, cursos=cursos, grupos=grupos, grupos_prof=grupos_prof)


@app.route('/profesores/<int:id>/baja', methods=['POST'])
@login_required
def marcar_baja(id):
    if not current_user.es_admin:
        flash('Solo administradores.', 'danger')
        return redirect(url_for('profesores'))
    p = Profesor.query.get_or_404(id)
    p.de_baja = True
    p.fecha_baja = date.today()
    ausencia = Ausencia(
        profesor_id=p.id,
        fecha_inicio=date.today(),
        motivo=request.form.get('motivo', 'Baja médica'),
        es_baja=True,
        notas=request.form.get('notas', ''),
        reportada_por=current_user.id
    )
    db.session.add(ausencia)
    db.session.commit()
    flash(f'{p.nombre} marcado de baja.', 'warning')
    return redirect(url_for('profesores'))


@app.route('/profesores/<int:id>/alta', methods=['POST'])
@login_required
def marcar_alta(id):
    if not current_user.es_admin:
        flash('Solo administradores.', 'danger')
        return redirect(url_for('profesores'))
    p = Profesor.query.get_or_404(id)
    p.de_baja = False
    p.fecha_vuelta = date.today()
    ausencia = Ausencia.query.filter_by(profesor_id=p.id, es_baja=True, fecha_fin=None).first()
    if ausencia:
        ausencia.fecha_fin = date.today()
    db.session.commit()
    flash(f'{p.nombre} ha vuelto de baja.', 'success')
    return redirect(url_for('profesores'))


# ───────────────────────────── HORARIO PERSONAL ─────────────────────────────

@app.route('/mi-horario')
@login_required
def mi_horario():
    horario = {}
    for dia in DIAS_SEMANA:
        horario[dia] = {}
        for franja in FRANJAS:
            h = HorarioProfesor.query.filter_by(
                profesor_id=current_user.id, dia=dia, franja=franja
            ).first()
            ind = Indisponibilidad.query.filter_by(
                profesor_id=current_user.id, dia=dia, franja=franja, activa=True, recurrente=True
            ).first()
            horario[dia][franja] = {
                'tiene_clase': h.tiene_clase if h else False,
                'asignatura': h.asignatura if h else '',
                'indisponible': ind is not None,
                'motivo_ind': ind.motivo if ind else ''
            }
    return render_template('mi_horario.html', horario=horario, dias=DIAS_SEMANA, franjas=FRANJAS)


@app.route('/mi-horario/guardar', methods=['POST'])
@login_required
def guardar_horario():
    data = request.get_json()
    for dia in DIAS_SEMANA:
        for franja in FRANJAS:
            key = f'{dia}_{franja}'
            celda = data.get(key, {})
            h = HorarioProfesor.query.filter_by(
                profesor_id=current_user.id, dia=dia, franja=franja
            ).first()
            if not h:
                h = HorarioProfesor(profesor_id=current_user.id, dia=dia, franja=franja)
                db.session.add(h)
            h.tiene_clase = celda.get('tiene_clase', False)
            h.asignatura = celda.get('asignatura', '')

            ind = Indisponibilidad.query.filter_by(
                profesor_id=current_user.id, dia=dia, franja=franja, recurrente=True
            ).first()
            if celda.get('indisponible'):
                if not ind:
                    ind = Indisponibilidad(
                        profesor_id=current_user.id, dia=dia, franja=franja, recurrente=True
                    )
                    db.session.add(ind)
                ind.motivo = celda.get('motivo_ind', '')
                ind.activa = True
            elif ind:
                ind.activa = False
    db.session.commit()
    return jsonify({'ok': True})


# ───────────────────────────── AUSENCIAS ─────────────────────────────

@app.route('/ausencias')
@login_required
def ausencias():
    if current_user.es_admin:
        lista = Ausencia.query.order_by(Ausencia.creada.desc()).all()
    else:
        lista = Ausencia.query.filter_by(profesor_id=current_user.id).order_by(Ausencia.creada.desc()).all()
    return render_template('ausencias.html', ausencias=lista)


def crear_guardias_desde_ausencia(ausencia):
    """Crea y asigna guardias para una ausencia. Devuelve lista de guardias creadas."""
    profesor = Profesor.query.get(ausencia.profesor_id)
    fecha = ausencia.fecha_inicio
    guardias_creadas = []

    # Determinar qué franjas cubrir
    if ausencia.franjas_afectadas == 'todas':
        franjas = FRANJAS
    else:
        franjas = json.loads(ausencia.franjas_afectadas)

    dia_semana = DIAS_SEMANA[fecha.weekday()] if fecha.weekday() < 5 else None
    if not dia_semana:
        return []

    for franja in franjas:
        candidato = buscar_profesor_para_guardia(
            dia_semana, franja, profesor.id,
            aula=profesor.aula_tutoria,
            etapa_ausente=profesor.etapa,
            fecha=fecha
        )
        guardia = Guardia(
            fecha=fecha,
            dia_semana=dia_semana,
            franja=franja,
            profesor_ausente_id=profesor.id,
            motivo_ausencia=ausencia.motivo,
            estado='pendiente' if candidato else 'sin_cobertura'
        )
        if candidato:
            guardia.profesor_asignado_id = candidato.id
        db.session.add(guardia)
        guardias_creadas.append((guardia, candidato))

    db.session.commit()

    for guardia, candidato in guardias_creadas:
        if candidato:
            enviar_email_guardia(guardia)

    return guardias_creadas


@app.route('/ausencias/nueva', methods=['GET', 'POST'])
@login_required
def nueva_ausencia():
    cfg = get_config()
    if request.method == 'POST':
        fecha_inicio = datetime.strptime(request.form.get('fecha_inicio'), '%Y-%m-%d').date()
        fecha_fin_str = request.form.get('fecha_fin', '')
        fecha_fin = datetime.strptime(fecha_fin_str, '%Y-%m-%d').date() if fecha_fin_str else None
        es_baja = bool(request.form.get('es_baja'))

        profesor_id = int(request.form.get('profesor_id', current_user.id)) if current_user.es_admin else current_user.id

        franjas_sel = request.form.getlist('franjas')
        franjas_afectadas = json.dumps(franjas_sel) if franjas_sel else 'todas'

        modo = cfg.modo_guardias
        aprobada = (modo == 'automatico') or current_user.es_admin

        ausencia = Ausencia(
            profesor_id=profesor_id,
            fecha_inicio=fecha_inicio,
            fecha_fin=fecha_fin,
            motivo=request.form.get('motivo', '').strip(),
            es_baja=es_baja,
            notas=request.form.get('notas', '').strip(),
            reportada_por=current_user.id,
            franjas_afectadas=franjas_afectadas,
            aprobada=aprobada
        )
        db.session.add(ausencia)

        if es_baja:
            p = Profesor.query.get(profesor_id)
            p.de_baja = True
            p.fecha_baja = fecha_inicio

        db.session.commit()

        if aprobada and not es_baja:
            guardias = crear_guardias_desde_ausencia(ausencia)
            asignadas = sum(1 for _, c in guardias if c)
            sin_cobertura = len(guardias) - asignadas
            msg = f'Ausencia registrada. {asignadas} guardia(s) asignada(s) automáticamente.'
            if sin_cobertura:
                msg += f' {sin_cobertura} franja(s) sin cobertura disponible.'
            flash(msg, 'success' if not sin_cobertura else 'warning')
        elif not aprobada:
            flash('Ausencia registrada. El administrador la revisará y asignará la guardia.', 'info')
            # Notificar al admin por email si está configurado
            _notificar_admin_ausencia_pendiente(ausencia)
        else:
            flash('Baja registrada correctamente.', 'success')

        return redirect(url_for('ausencias'))

    profesores_lista = Profesor.query.filter_by(activo=True).order_by(Profesor.nombre).all() if current_user.es_admin else []
    return render_template('form_ausencia.html', profesores=profesores_lista, franjas=FRANJAS, cfg=cfg)


def _notificar_admin_ausencia_pendiente(ausencia):
    if not mail_configurado():
        return
    try:
        cfg = get_mail_config()
        init_mail(cfg)
        admins = Profesor.query.filter_by(es_admin=True, activo=True).all()
        profesor = Profesor.query.get(ausencia.profesor_id)
        for admin in admins:
            url = url_for('aprobar_ausencia', id=ausencia.id, _external=True)
            msg = Message(
                subject=f'Ausencia pendiente de aprobación — {profesor.nombre}',
                recipients=[admin.email]
            )
            msg.html = render_template('email_ausencia_pendiente.html',
                admin=admin, profesor=profesor, ausencia=ausencia, url=url)
            mail.send(msg)
    except Exception as e:
        print(f'Error notificando admin: {e}')


@app.route('/ausencias/<int:id>/aprobar', methods=['POST'])
@login_required
def aprobar_ausencia(id):
    if not current_user.es_admin:
        flash('Solo administradores.', 'danger')
        return redirect(url_for('ausencias'))
    ausencia = Ausencia.query.get_or_404(id)
    ausencia.aprobada = True
    db.session.commit()
    guardias = crear_guardias_desde_ausencia(ausencia)
    asignadas = sum(1 for _, c in guardias if c)
    sin_cobertura = len(guardias) - asignadas
    msg = f'{asignadas} guardia(s) asignada(s).'
    if sin_cobertura:
        msg += f' {sin_cobertura} franja(s) sin cobertura.'
    flash(msg, 'success' if not sin_cobertura else 'warning')
    return redirect(url_for('ausencias'))


# ───────────────────────────── GUARDIAS ─────────────────────────────

@app.route('/guardias')
@login_required
def guardias():
    if current_user.es_admin:
        lista = Guardia.query.order_by(Guardia.fecha.desc(), Guardia.franja).all()
    else:
        lista = Guardia.query.filter(
            (Guardia.profesor_asignado_id == current_user.id) |
            (Guardia.profesor_ausente_id == current_user.id)
        ).order_by(Guardia.fecha.desc()).all()
    return render_template('guardias.html', guardias=lista)


@app.route('/guardias/nueva', methods=['GET', 'POST'])
@login_required
def nueva_guardia():
    if request.method == 'POST':
        fecha = datetime.strptime(request.form.get('fecha'), '%Y-%m-%d').date()
        dia_semana = DIAS_SEMANA[fecha.weekday()] if fecha.weekday() < 5 else 'Lunes'
        franja = request.form.get('franja')
        profesor_ausente_id = int(request.form.get('profesor_ausente_id'))

        aula = request.form.get('aula', '').strip()
        ausente = Profesor.query.get(profesor_ausente_id)
        etapa_ausente = ausente.etapa if ausente else None

        candidato = buscar_profesor_para_guardia(
            dia_semana, franja, profesor_ausente_id,
            aula=aula, etapa_ausente=etapa_ausente, fecha=fecha
        )

        guardia = Guardia(
            fecha=fecha,
            dia_semana=dia_semana,
            franja=franja,
            profesor_ausente_id=profesor_ausente_id,
            motivo_ausencia=request.form.get('motivo_ausencia', '').strip(),
            aula=aula,
            notas=request.form.get('notas', '').strip()
        )

        if candidato:
            guardia.profesor_asignado_id = candidato.id
            guardia.estado = 'pendiente'
            db.session.add(guardia)
            db.session.commit()
            email_ok = enviar_email_guardia(guardia)
            if email_ok:
                flash(f'Guardia asignada a {candidato.nombre} y notificado por email.', 'success')
            else:
                flash(f'Guardia asignada a {candidato.nombre}. (Email no configurado aún)', 'warning')
        else:
            guardia.estado = 'sin_cobertura'
            db.session.add(guardia)
            db.session.commit()
            flash('No hay ningún profesor disponible para esa franja. Guardia sin cobertura.', 'danger')

        return redirect(url_for('guardias'))

    profesores_lista = Profesor.query.filter_by(activo=True, de_baja=False).order_by(Profesor.nombre).all()
    return render_template('form_guardia.html', profesores=profesores_lista, franjas=FRANJAS)


@app.route('/guardia/<int:id>/completar', methods=['POST'])
@login_required
def completar_guardia(id):
    g = Guardia.query.get_or_404(id)
    g.completada = True
    g.estado = 'completada'
    db.session.commit()
    flash('Guardia marcada como completada.', 'success')
    return redirect(url_for('guardias'))


@app.route('/accion-guardia/<token>')
def accion_guardia(token):
    try:
        data = serializer.loads(token, salt='guardia-action', max_age=86400)
        accion, guardia_id = data.split(':')
        guardia = Guardia.query.get_or_404(int(guardia_id))
        if accion == 'confirmar':
            guardia.estado = 'confirmada'
            db.session.commit()
            return render_template('accion_guardia.html', accion='confirmada', guardia=guardia)
        elif accion == 'rechazar':
            guardia.estado = 'rechazada'
            candidato_anterior_id = guardia.profesor_asignado_id
            nuevo = buscar_profesor_para_guardia(
                guardia.dia_semana, guardia.franja, candidato_anterior_id,
                aula=guardia.aula,
                etapa_ausente=guardia.profesor_ausente.etapa if guardia.profesor_ausente else None,
                fecha=guardia.fecha
            )
            if nuevo and nuevo.id != candidato_anterior_id:
                guardia.profesor_asignado_id = nuevo.id
                guardia.estado = 'pendiente'
                db.session.commit()
                enviar_email_guardia(guardia)
                return render_template('accion_guardia.html', accion='rechazada_reasignada', guardia=guardia, nuevo=nuevo)
            else:
                guardia.estado = 'sin_cobertura'
                db.session.commit()
                return render_template('accion_guardia.html', accion='sin_cobertura', guardia=guardia)
    except Exception:
        return render_template('accion_guardia.html', accion='error', guardia=None)


# ───────────────────────────── IMPORTAR HORARIOS ─────────────────────────────

@app.route('/horarios/importar', methods=['GET', 'POST'])
@login_required
def importar_horarios():
    if not current_user.es_admin:
        flash('Solo administradores.', 'danger')
        return redirect(url_for('dashboard'))

    if request.method == 'POST':
        f = request.files.get('archivo')
        if not f or not f.filename.endswith(('.xlsx', '.xls')):
            flash('Sube un fichero Excel (.xlsx).', 'danger')
            return redirect(url_for('importar_horarios'))

        try:
            wb = openpyxl.load_workbook(io.BytesIO(f.read()))
            ws = wb.active
            filas = list(ws.iter_rows(values_only=True))
            if not filas:
                flash('El fichero está vacío.', 'danger')
                return redirect(url_for('importar_horarios'))

            # Cabecera esperada: Profesor | Dia | Franja | Tipo | Asignatura
            errores = []
            importados = 0
            for i, fila in enumerate(filas[1:], start=2):  # salta cabecera
                if not any(fila):
                    continue
                nombre_prof = str(fila[0]).strip() if fila[0] else ''
                dia = str(fila[1]).strip() if fila[1] else ''
                franja = str(fila[2]).strip() if fila[2] else ''
                tipo = str(fila[3]).strip().lower() if fila[3] else ''
                asignatura = str(fila[4]).strip() if len(fila) > 4 and fila[4] else ''

                if not nombre_prof or not dia or not franja:
                    errores.append(f'Fila {i}: datos incompletos')
                    continue
                if dia not in DIAS_SEMANA:
                    errores.append(f'Fila {i}: día "{dia}" no reconocido')
                    continue
                if franja not in FRANJAS:
                    errores.append(f'Fila {i}: franja "{franja}" no reconocida')
                    continue

                # Buscar profesor por nombre (parcial, sin distinción mayúsculas)
                profesor = Profesor.query.filter(
                    Profesor.nombre.ilike(f'%{nombre_prof}%')
                ).first()
                if not profesor:
                    errores.append(f'Fila {i}: profesor "{nombre_prof}" no encontrado')
                    continue

                tiene_clase = tipo in ('clase', 'sí', 'si', 'yes', '1', 'true')
                es_indisponible = tipo in ('reunión', 'reunion', 'no disponible', 'ocupado', 'ind')

                h = HorarioProfesor.query.filter_by(
                    profesor_id=profesor.id, dia=dia, franja=franja
                ).first()
                if not h:
                    h = HorarioProfesor(profesor_id=profesor.id, dia=dia, franja=franja)
                    db.session.add(h)
                h.tiene_clase = tiene_clase
                h.asignatura = asignatura if tiene_clase else ''

                if es_indisponible:
                    ind = Indisponibilidad.query.filter_by(
                        profesor_id=profesor.id, dia=dia, franja=franja, recurrente=True
                    ).first()
                    if not ind:
                        ind = Indisponibilidad(
                            profesor_id=profesor.id, dia=dia, franja=franja, recurrente=True
                        )
                        db.session.add(ind)
                    ind.motivo = asignatura or 'Ocupado'
                    ind.activa = True

                importados += 1

            db.session.commit()
            msg = f'Importados {importados} registros correctamente.'
            if errores:
                msg += f' {len(errores)} filas con errores: ' + '; '.join(errores[:5])
                if len(errores) > 5:
                    msg += f' y {len(errores)-5} más.'
                flash(msg, 'warning')
            else:
                flash(msg, 'success')

        except Exception as e:
            flash(f'Error procesando el fichero: {e}', 'danger')

        return redirect(url_for('importar_horarios'))

    profesores_lista = Profesor.query.filter_by(activo=True).order_by(Profesor.nombre).all()
    return render_template('importar_horarios.html', profesores=profesores_lista, dias=DIAS_SEMANA, franjas=FRANJAS)


@app.route('/horarios/plantilla')
@login_required
def descargar_plantilla():
    if not current_user.es_admin:
        flash('Solo administradores.', 'danger')
        return redirect(url_for('dashboard'))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Horarios'

    # Cabecera
    cabecera = ['Profesor', 'Dia', 'Franja', 'Tipo', 'Asignatura']
    for col, titulo in enumerate(cabecera, 1):
        ws.cell(row=1, column=col, value=titulo)
        ws.cell(row=1, column=col).font = openpyxl.styles.Font(bold=True)

    # Filas de ejemplo
    ejemplos = [
        ['García López, María', 'Lunes', '1ª hora', 'clase', 'Matemáticas'],
        ['García López, María', 'Lunes', '2ª hora', 'libre', ''],
        ['García López, María', 'Martes', 'Recreo', 'reunion', 'Reunión de departamento'],
        ['Fernández Ruiz, Juan', 'Lunes', '1ª hora', 'libre', ''],
    ]
    for row, fila in enumerate(ejemplos, 2):
        for col, val in enumerate(fila, 1):
            ws.cell(row=row, column=col, value=val)

    # Hoja de ayuda
    ws2 = wb.create_sheet('Instrucciones')
    instrucciones = [
        ['INSTRUCCIONES DE USO'],
        [''],
        ['Columna "Tipo" — valores válidos:'],
        ['  clase       → el profesor tiene clase en esa franja (no puede hacer guardia)'],
        ['  libre        → franja libre (puede hacer guardia)'],
        ['  reunion     → ocupado por reunión u otro motivo (no puede hacer guardia)'],
        [''],
        ['Columna "Dia" — valores exactos:'],
        ['  Lunes, Martes, Miércoles, Jueves, Viernes'],
        [''],
        ['Columna "Franja" — valores exactos:'],
        ['  1ª hora, 2ª hora, 3ª hora, Recreo, 4ª hora, 5ª hora, 6ª hora'],
        [''],
        ['El nombre del profesor debe coincidir (al menos en parte) con el nombre en la aplicación.'],
    ]
    for row, linea in enumerate(instrucciones, 1):
        ws2.cell(row=row, column=1, value=linea[0])

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 30

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    from flask import send_file
    return send_file(output, download_name='plantilla_horarios.xlsx',
                     as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ───────────────────────────── ESTADÍSTICAS ─────────────────────────────

@app.route('/estadisticas')
@login_required
def estadisticas():
    profesores_lista = Profesor.query.filter_by(activo=True).order_by(Profesor.nombre).all()
    hoy = date.today()
    guardias_por_mes = []
    for mes in range(1, hoy.month + 1):
        count = Guardia.query.filter(
            db.extract('month', Guardia.fecha) == mes,
            db.extract('year', Guardia.fecha) == hoy.year
        ).count()
        guardias_por_mes.append({'mes': mes, 'total': count})
    return render_template('estadisticas.html',
        profesores=profesores_lista,
        guardias_por_mes=guardias_por_mes)


# ───────────────────────────── CONFIGURACIÓN EMAIL ─────────────────────────────

@app.route('/configuracion', methods=['GET', 'POST'])
@login_required
def configuracion():
    if not current_user.es_admin:
        flash('Solo administradores.', 'danger')
        return redirect(url_for('dashboard'))
    cfg = get_config()
    if request.method == 'POST':
        cfg.mail_username = request.form.get('mail_username', '').strip()
        cfg.mail_password = request.form.get('mail_password', '').strip()
        cfg.activo = bool(request.form.get('activo'))
        api_key = request.form.get('anthropic_api_key', '').strip()
        if api_key:
            cfg.anthropic_api_key = api_key
        db.session.commit()
        flash('Configuración guardada.', 'success')
    return render_template('configuracion.html', cfg=cfg)


@app.route('/configuracion/modo', methods=['POST'])
@login_required
def cambiar_modo():
    if not current_user.es_admin:
        flash('Solo administradores.', 'danger')
        return redirect(url_for('configuracion'))
    cfg = get_config()
    cfg.modo_guardias = request.form.get('modo_guardias', 'manual')
    db.session.commit()
    modo_texto = 'Automático' if cfg.modo_guardias == 'automatico' else 'Manual'
    flash(f'Modo cambiado a: {modo_texto}.', 'success')
    return redirect(url_for('configuracion'))


# ───────────────────────────── TUTORÍAS ─────────────────────────────

@app.route('/tutorias', methods=['GET', 'POST'])
@login_required
def tutorias():
    if not current_user.es_admin:
        flash('Solo administradores.', 'danger')
        return redirect(url_for('dashboard'))

    if request.method == 'POST':
        accion = request.form.get('accion')

        if accion == 'add_curso':
            nombre = request.form.get('nombre', '').strip()
            if nombre:
                if not Curso.query.filter_by(nombre=nombre).first():
                    orden = Curso.query.count()
                    db.session.add(Curso(
                        nombre=nombre, orden=orden,
                        etapa=request.form.get('etapa', '').strip() or None,
                        aula_cerrada=bool(request.form.get('aula_cerrada'))
                    ))
                    db.session.commit()
                    flash(f'Curso "{nombre}" añadido.', 'success')
                else:
                    flash(f'El curso "{nombre}" ya existe.', 'warning')

        elif accion == 'edit_curso':
            curso_id = int(request.form.get('curso_id', 0))
            curso = Curso.query.get_or_404(curso_id)
            curso.etapa = request.form.get('etapa', '').strip() or None
            curso.aula_cerrada = bool(request.form.get('aula_cerrada'))
            db.session.commit()
            flash('Curso actualizado.', 'success')

        elif accion == 'delete_curso':
            curso_id = int(request.form.get('curso_id', 0))
            curso = Curso.query.get_or_404(curso_id)
            # Desasignar tutores que tenían este curso
            Profesor.query.filter_by(aula_tutoria=curso.nombre).update({'aula_tutoria': None})
            db.session.delete(curso)
            db.session.commit()
            flash(f'Curso eliminado.', 'success')

        elif accion == 'assign_tutor':
            curso_id = int(request.form.get('curso_id', 0))
            profesor_id = request.form.get('profesor_id', '')
            curso = Curso.query.get_or_404(curso_id)
            if profesor_id:
                p = Profesor.query.get_or_404(int(profesor_id))
                p.aula_tutoria = curso.nombre
                db.session.commit()
                flash('Tutor añadido.', 'success')

        elif accion == 'remove_tutor':
            profesor_id = int(request.form.get('profesor_id', 0))
            p = Profesor.query.get_or_404(profesor_id)
            p.aula_tutoria = None
            db.session.commit()
            flash('Tutor eliminado.', 'success')

        return redirect(url_for('tutorias'))

    cursos = Curso.query.order_by(Curso.orden, Curso.nombre).all()
    profesores_lista = Profesor.query.filter_by(activo=True, de_baja=False).order_by(Profesor.nombre).all()
    tutores = {
        c.nombre: Profesor.query.filter_by(aula_tutoria=c.nombre).all()
        for c in cursos
    }
    return render_template('tutorias.html',
        cursos=cursos, profesores=profesores_lista, tutores=tutores)


# ───────────────────────────── GRUPOS DE TRABAJO ─────────────────────────────

@app.route('/grupos-trabajo', methods=['GET', 'POST'])
@login_required
def grupos_trabajo():
    if not current_user.es_admin:
        return redirect(url_for('dashboard'))
    if request.method == 'POST':
        accion = request.form.get('accion')
        if accion == 'add_grupo':
            nombre = request.form.get('nombre', '').strip()
            if nombre and not GrupoTrabajo.query.filter_by(nombre=nombre).first():
                db.session.add(GrupoTrabajo(nombre=nombre))
                db.session.commit()
                flash(f'Grupo «{nombre}» creado.', 'success')
            else:
                flash('Nombre vacío o ya existe.', 'danger')
        elif accion == 'delete_grupo':
            g = GrupoTrabajo.query.get_or_404(int(request.form.get('grupo_id')))
            db.session.delete(g)
            db.session.commit()
            flash('Grupo eliminado.', 'success')
        elif accion == 'guardar_miembros':
            grupo_id = int(request.form.get('grupo_id'))
            g = GrupoTrabajo.query.get_or_404(grupo_id)
            # Delete existing and re-create from form
            ProfesorGrupo.query.filter_by(grupo_id=grupo_id).delete()
            for p in Profesor.query.filter_by(activo=True).all():
                key = f'horas_{p.id}'
                if key in request.form and request.form.get(f'miembro_{p.id}'):
                    horas = float(request.form.get(key) or 0)
                    if horas > 0:
                        db.session.add(ProfesorGrupo(
                            profesor_id=p.id, grupo_id=grupo_id, horas_semanales=horas))
            db.session.commit()
            flash('Miembros actualizados.', 'success')
        return redirect(url_for('grupos_trabajo'))
    grupos = GrupoTrabajo.query.order_by(GrupoTrabajo.nombre).all()
    profesores = Profesor.query.filter_by(activo=True, de_baja=False).order_by(Profesor.nombre).all()
    miembros = {g.id: {m.profesor_id: m.horas_semanales for m in g.miembros} for g in grupos}
    return render_template('grupos_trabajo.html', grupos=grupos, profesores=profesores, miembros=miembros)


# ───────────────────────────── SEED DATOS ─────────────────────────────

_SEED_HORARIO = {
    'HH1': {'Religión':2,'Inglés':2,'CRR':12,'CEA':6,'DEE':3},
    'HH2': {'Religión':2,'Inglés':2,'CRR':13,'CEA':5,'DEE':3},
    'HH3': {'Religión':2,'Inglés':2,'CRR':12,'CEA':6,'DEE':3},
    'LH1': {'Euskera':4,'Inglés':3,'Gimnasia':2,'Lengua':3.5,'Lectura Lengua':0.5,
            'Música':1,'Inguru':3,'Religión':2,'Mate':4,'Tutoretza':1,'Plastika':1},
    'LH2': {'Euskera':3,'Inglés':3,'Gimnasia':3,'Lengua':3.5,'Lectura Lengua':0.5,
            'Música':1,'Inguru':3,'Religión':2,'Mate':4,'Tutoretza':1,'Plastika':1},
    'LH3': {'Euskera':4,'Mate':4,'Inglés':3,'Gimnasia':2,'Música':1,'Inguru':4,
            'Tutoretza':1,'Lengua':3,'Religión':2,'Plastika':1},
    'LH4': {'Euskera':4,'Inglés':3,'Gimnasia':2,'Lengua':4,'Música':1,'Inguru':3,
            'Religión':2,'Mate':4,'Tutoretza':1,'Plastika':1},
    'LH5': {'Euskera':3.5,'Lectura Euskera':0.5,'Inglés':2,'Gimnasia':2,'Lengua':4,
            'Música':1,'Inguru':4,'Religión':2,'Mate':3,'Francés':1,'Tutoretza':1,'Plastika':1},
    'LH6': {'Euskera':2.5,'Lectura Euskera':0.5,'Inglés':3,'Gimnasia':2,'Lengua':3,
            'Música':1,'Inguru':3,'Religión':2,'Valores':2,'Mate':3,'Francés':1,
            'Tutoretza':1,'Plastika':1},
}

@app.route('/admin/autoasignar-asignaturas')
@login_required
def autoasignar_asignaturas_todos():
    if not current_user.es_admin:
        return redirect(url_for('dashboard'))
    profesores = Profesor.query.filter_by(activo=True, es_admin=False).all()
    for p in profesores:
        _auto_asignar_asignaturas(p)
    db.session.commit()
    flash(f'Asignaturas actualizadas para {len(profesores)} profesores.', 'success')
    return redirect(url_for('horarios_construccion', tab='profesores'))


@app.route('/admin/importar-datos-iniciales')
@login_required
def importar_datos_iniciales():
    if not current_user.es_admin:
        return redirect(url_for('dashboard'))

    creados = {'cursos': 0, 'asignaturas': 0, 'asignaciones': 0}

    for i, nombre in enumerate(['HH1','HH2','HH3']):
        if not Curso.query.filter_by(nombre=nombre).first():
            db.session.add(Curso(nombre=nombre, etapa='Haur Hezkuntza', orden=i+1))
            creados['cursos'] += 1
    for i, nombre in enumerate(['LH1','LH2','LH3','LH4','LH5','LH6']):
        if not Curso.query.filter_by(nombre=nombre).first():
            db.session.add(Curso(nombre=nombre, etapa='Lehen Hezkuntza', orden=10+i))
            creados['cursos'] += 1
    db.session.flush()

    # Determinar etapa de cada asignatura: None si aparece en varias etapas
    _etapa_asig = {}
    for curso_nombre, asigs in _SEED_HORARIO.items():
        etapa = 'Haur Hezkuntza' if curso_nombre.startswith('HH') else 'Lehen Hezkuntza'
        for nombre in asigs:
            if nombre in _etapa_asig:
                if _etapa_asig[nombre] != etapa:
                    _etapa_asig[nombre] = None  # aparece en ambas etapas → sin restricción
            else:
                _etapa_asig[nombre] = etapa

    for nombre, etapa in _etapa_asig.items():
        asig = Asignatura.query.filter_by(nombre=nombre).first()
        if not asig:
            db.session.add(Asignatura(nombre=nombre, etapa=etapa))
            creados['asignaturas'] += 1
        else:
            if not asig.etapa:
                asig.etapa = etapa
    db.session.flush()

    # Upsert todas las asignaciones con las horas correctas (siempre actualiza)
    for curso_nombre, asigs in _SEED_HORARIO.items():
        curso = Curso.query.filter_by(nombre=curso_nombre).first()
        if not curso:
            continue
        for asig_nombre, horas in asigs.items():
            asig = Asignatura.query.filter_by(nombre=asig_nombre).first()
            if not asig:
                continue
            existing = CursoAsignatura.query.filter_by(
                curso_id=curso.id, asignatura_id=asig.id).first()
            if existing:
                existing.horas_semanales = float(horas)
            else:
                db.session.add(CursoAsignatura(
                    curso_id=curso.id, asignatura_id=asig.id, horas_semanales=float(horas)))
            creados['asignaciones'] += 1
    db.session.commit()
    flash(f'Importación completada: {creados["cursos"]} cursos, '
          f'{creados["asignaturas"]} asignaturas, {creados["asignaciones"]} asignaciones actualizadas.', 'success')
    return redirect(url_for('horarios_construccion', tab='asignaturas'))


# ───────────────────────────── INIT DB ─────────────────────────────

# Estado compartido para generación en background
_gen_estado = {
    'activo': False,
    'intentos': 0,
    'mejor_fallos': None,
    'error': None,
    'avisos': [],
    'listo': False,
}
_gen_lock = threading.Lock()


def generar_horario_automatico():
    import random

    HorarioAsignacion.query.filter_by(es_manual=False).delete()
    # Limpiar complementarias automáticas antes de regenerar
    try:
        SlotComplementaria.query.filter_by(es_manual=False).delete()
    except Exception:
        SlotComplementaria.query.delete()
    db.session.flush()

    # Pre-cargar celdas puestas a mano (no se tocan durante la generación)
    manual_asigs = HorarioAsignacion.query.filter_by(es_manual=True).all()
    manual_slots_prof = defaultdict(set)   # prof_id -> {(dia, franja)}
    manual_slots_curso = defaultdict(set)  # curso_id -> {(dia, franja)}
    manual_covered = []                    # [(curso_id, asig_id, prof_id), ...]
    for ma in manual_asigs:
        manual_slots_prof[ma.profesor_id].add((ma.dia, ma.franja))
        manual_slots_curso[ma.curso_id].add((ma.dia, ma.franja))
        manual_covered.append((ma.curso_id, ma.asignatura_id, ma.profesor_id))
        if ma.profesor2_id:
            manual_slots_prof[ma.profesor2_id].add((ma.dia, ma.franja))

    # Bloquear también los slots de grupos/complementarias manuales (no colocar clases encima)
    try:
        manual_slots_comp = SlotComplementaria.query.filter_by(es_manual=True).all()
    except Exception:
        manual_slots_comp = []
    for sc in manual_slots_comp:
        manual_slots_prof[sc.profesor_id].add((sc.dia, sc.franja))

    cursos_map = {c.id: c for c in Curso.query.all()}
    asig_map = {a.id: a for a in Asignatura.query.all()}
    profesores_activos = Profesor.query.filter_by(activo=True, es_admin=False).all()

    # Cargar reglas de asignatura
    reglas_max_dia = {}
    reglas_consecutivas = {}
    reglas_fijar = {}
    regla_tutor_primera = False
    regla_unico_prof = None  # None | 'dura' | 'blanda'

    # Reglas generales por clase (etapa): tutor imparte asignatura en su clase
    # etapa -> asig_id -> dureza ('dura'|'blanda')
    reglas_tutor_etapa_dura = defaultdict(set)   # etapa -> set(asig_id)
    reglas_tutor_etapa_blanda = defaultdict(set) # etapa -> set(asig_id)

    # Cargar reglas de profesor
    rp_excluir_curso = defaultdict(set)        # curso_id -> profs excluidos (dura)
    rp_fijar_curso = defaultdict(set)          # curso_id -> profs forzados (dura)
    rp_excluir_franja = defaultdict(set)       # (dia,franja) -> profs excluidos (dura)
    rp_fijar_asignatura = defaultdict(set)     # asig_id -> profs forzados (dura)
    rp_fijar_curso_asig = defaultdict(set)     # (curso_id, asig_id) -> profs forzados (dura)
    rp_evitar_curso = defaultdict(set)         # curso_id -> profs a evitar (blanda)
    rp_preferir_curso = defaultdict(set)       # curso_id -> profs preferidos (blanda)
    rp_evitar_franja = defaultdict(set)        # (dia,franja) -> profs a evitar (blanda)
    rp_preferir_asignatura = defaultdict(set)  # asig_id -> profs preferidos (blanda)
    rp_horas_etapa = defaultdict(dict)         # prof_id -> {etapa: max_horas}
    rp_min_horas = {}                         # prof_id -> min horas guardia
    rp_max_horas = {}                         # prof_id -> max horas guardia

    for r in ReglaHorario.query.all():
        if r.tipo == 'max_dia' and r.asignatura_id:
            reglas_max_dia[r.asignatura_id] = r.valor
        elif r.tipo == 'consecutivas' and r.asignatura_id:
            reglas_consecutivas[r.asignatura_id] = r.valor
        elif r.tipo == 'fijar_franja' and r.asignatura_id and r.dia and r.franja:
            reglas_fijar[r.asignatura_id] = (r.dia, r.franja)
        elif r.tipo == 'tutor_primera':
            regla_tutor_primera = True
        elif r.tipo == 'asig_unico_prof':
            regla_unico_prof = r.dureza  # 'dura' | 'blanda'
        elif r.tipo == 'tutor_clase_etapa' and r.etapa and r.asignatura_id:
            if r.dureza == 'dura':
                reglas_tutor_etapa_dura[r.etapa].add(r.asignatura_id)
            else:
                reglas_tutor_etapa_blanda[r.etapa].add(r.asignatura_id)
        elif r.profesor_id:
            pid = r.profesor_id
            if r.tipo == 'prof_excluir_curso' and r.curso_id_regla:
                rp_excluir_curso[r.curso_id_regla].add(pid)
            elif r.tipo == 'prof_fijar_curso' and r.curso_id_regla:
                rp_fijar_curso[r.curso_id_regla].add(pid)
            elif r.tipo == 'prof_excluir_franja' and r.dia and r.franja:
                rp_excluir_franja[(r.dia, r.franja)].add(pid)
            elif r.tipo == 'prof_fijar_asignatura' and r.asignatura_id:
                rp_fijar_asignatura[r.asignatura_id].add(pid)
            elif r.tipo == 'prof_fijar_curso_asignatura' and r.curso_id_regla and r.asignatura_id:
                rp_fijar_curso_asig[(r.curso_id_regla, r.asignatura_id)].add(pid)
            elif r.tipo == 'prof_evitar_curso' and r.curso_id_regla:
                rp_evitar_curso[r.curso_id_regla].add(pid)
            elif r.tipo == 'prof_preferir_curso' and r.curso_id_regla:
                rp_preferir_curso[r.curso_id_regla].add(pid)
            elif r.tipo == 'prof_evitar_franja' and r.dia and r.franja:
                rp_evitar_franja[(r.dia, r.franja)].add(pid)
            elif r.tipo == 'prof_preferir_asignatura' and r.asignatura_id:
                rp_preferir_asignatura[r.asignatura_id].add(pid)
            elif r.tipo == 'prof_min_horas':
                rp_min_horas[pid] = r.valor
            elif r.tipo == 'prof_max_horas':
                rp_max_horas[pid] = r.valor
            elif r.tipo == 'prof_horas_etapa' and r.etapa:
                rp_horas_etapa[pid][r.etapa] = r.valor

    import math
    tasks = []
    half_tasks_per_curso = defaultdict(list)  # curso_id -> [asig_id, ...]

    for req in CursoAsignatura.query.all():
        curso = cursos_map.get(req.curso_id)
        if curso and curso.aula_cerrada:
            continue
        horas = req.horas_semanales or 0
        n_full = int(horas)
        has_half = round(horas % 1, 1) >= 0.4  # >= 0.5 con tolerancia float
        for _ in range(n_full):
            tasks.append((req.curso_id, req.asignatura_id))
        if has_half:
            half_tasks_per_curso[req.curso_id].append(req.asignatura_id)

    # Emparejar medias horas dentro del mismo curso
    paired_tasks = []  # [(curso_id, asig_id1, asig_id2)]
    for curso_id, halves in half_tasks_per_curso.items():
        while len(halves) >= 2:
            paired_tasks.append((curso_id, halves.pop(0), halves.pop(0)))

    prof_por_asig = defaultdict(list)
    for pa in ProfesorAsignatura.query.all():
        prof_por_asig[pa.asignatura_id].append(pa.profesor_id)

    prof_etapas = {p.id: parse_etapas(p.etapa) for p in profesores_activos}
    prof_tutoria = defaultdict(list)
    for p in profesores_activos:
        if p.aula_tutoria:
            prof_tutoria[p.aula_tutoria].append(p.id)
    curso_nombre = {c.id: c.nombre for c in cursos_map.values()}
    curso_etapa = {c.id: c.etapa for c in cursos_map.values()}

    def etapa_compatible(prof_id, asig_id):
        asig = asig_map.get(asig_id)
        if not asig or not asig.etapa:
            return True
        etapas_prof = prof_etapas.get(prof_id, [])
        if not etapas_prof:
            return True
        return asig.etapa in etapas_prof

    # ══════════════════════════════════════════════════════════════════════
    # FASE 1 — Asignación: ¿quién da qué en qué clase?
    # Más restringido primero. Reglas duras son absolutas — si no hay
    # candidato válido el slot queda sin asignar (no se viola ninguna regla).
    # ══════════════════════════════════════════════════════════════════════
    import math

    franjas_clase = [f for f in FRANJAS if f != 'Patio']
    primera_franja = franjas_clase[0] if franjas_clase else None

    prof_max = {
        p.id: float(p.horas_lectivas) if (p.horas_lectivas or 0) > 0 else (p.horas_max_semanales or 25)
        for p in profesores_activos
    }
    for pid, mx in rp_max_horas.items():
        prof_max[pid] = mx

    def eligible_for(curso_id, asig_id):
        """Profesores elegibles para (curso, asig) tras aplicar todas las reglas duras."""
        # prof_fijar_curso_asignatura es la regla más específica: override total.
        # No requiere que el profesor esté en ProfesorAsignatura.
        fij_ca = rp_fijar_curso_asig.get((curso_id, asig_id), set())
        if fij_ca:
            return fij_ca - rp_excluir_curso.get(curso_id, set())

        profs = {pid for pid in prof_por_asig.get(asig_id, [])
                 if etapa_compatible(pid, asig_id)}
        profs -= rp_excluir_curso.get(curso_id, set())
        fij_a = rp_fijar_asignatura.get(asig_id, set())
        if fij_a:
            profs &= fij_a
        fij_c = rp_fijar_curso.get(curso_id, set())
        if fij_c:
            profs &= fij_c
        etapa = curso_etapa.get(curso_id)
        if etapa and asig_id in reglas_tutor_etapa_dura.get(etapa, set()):
            tutores = set(prof_tutoria.get(curso_nombre.get(curso_id, ''), []))
            profs &= tutores
        return profs

    def sort_eligible(profs, curso_id, asig_id, load):
        """Ordena candidatos: tutores/preferidos primero, evitados al final.
        Dentro de cada tier, el que más horas le faltan para su objetivo va primero."""
        etapa = curso_etapa.get(curso_id)
        tutores = set(prof_tutoria.get(curso_nombre.get(curso_id, ''), []))
        preferir = rp_preferir_curso.get(curso_id, set()) | rp_preferir_asignatura.get(asig_id, set())
        evitar = rp_evitar_curso.get(curso_id, set())
        blanda_tutor = bool(etapa and asig_id in reglas_tutor_etapa_blanda.get(etapa, set()))
        def key(p):
            if blanda_tutor and p in tutores:
                tier = 0
            elif p in preferir:
                tier = 1
            elif p in evitar:
                tier = 3
            else:
                tier = 2
            # Mayor capacidad restante = más urgente = va primero (valor más bajo)
            remaining = prof_max.get(p, 25) - load[p]
            return (tier, -remaining)
        return sorted(profs, key=key)

    task_count = defaultdict(int)
    for curso_id, asig_id in tasks:
        task_count[(curso_id, asig_id)] += 1

    # Agrupar por número de candidatos elegibles (para priorizar más restringidos)
    _elig_cache = {(c, a): len(eligible_for(c, a)) for (c, a) in task_count}
    _pairs_by_n = defaultdict(list)
    for (c, a), cnt in task_count.items():
        _pairs_by_n[_elig_cache[(c, a)]].append(((c, a), cnt))

    def _run_fase1():
        """Un intento de asignación profesor→clase. Devuelve (normal, paired, unassignable, load, etapa_load, deficit)."""
        lload = defaultdict(int)                          # prof_id -> horas asignadas
        letapa = defaultdict(lambda: defaultdict(int))   # prof_id -> etapa -> horas

        def _quota_ok(pid, cid):
            quotas = rp_horas_etapa.get(pid)
            if not quotas:
                return True
            etapa = curso_etapa.get(cid)
            return not etapa or etapa not in quotas or letapa[pid][etapa] < quotas[etapa]

        def _etapa_add(pid, cid):
            etapa = curso_etapa.get(cid)
            if etapa:
                letapa[pid][etapa] += 1

        def _avail(profs, cid):
            return [p for p in profs if lload[p] < prof_max.get(p, 25) and _quota_ok(p, cid)]

        lnormal = []
        lpaired = []
        lunassig = []

        # Mezclar dentro de cada grupo de igual restricción para explorar ordenes distintos
        shuffled = []
        for n in sorted(_pairs_by_n):
            g = list(_pairs_by_n[n])
            random.shuffle(g)
            shuffled.extend(g)

        for (curso_id, asig_id), count in shuffled:
            eligible = eligible_for(curso_id, asig_id)
            fij_ca = rp_fijar_curso_asig.get((curso_id, asig_id), set())
            is_cotutor = len(fij_ca & eligible) > 1

            if is_cotutor:
                for _ in range(count):
                    avail = _avail(eligible, curso_id)
                    if not avail:
                        lunassig.append((curso_id, asig_id))
                        continue
                    prof = min(avail, key=lambda p: lload[p])
                    lnormal.append((curso_id, asig_id, prof))
                    lload[prof] += 1
                    _etapa_add(prof, curso_id)
            elif regla_unico_prof == 'dura':
                avail = sort_eligible(_avail(eligible, curso_id), curso_id, asig_id, lload)
                if not avail:
                    lunassig.extend([(curso_id, asig_id)] * count)
                    continue
                prof = avail[0]
                etapa = curso_etapa.get(curso_id)
                etapa_restante = (rp_horas_etapa.get(prof, {}).get(etapa, 9999)
                                  - letapa[prof].get(etapa, 0)) if etapa else 9999
                cap = min(prof_max.get(prof, 25) - lload[prof], etapa_restante)
                n = int(min(count, cap))
                for _ in range(n):
                    lnormal.append((curso_id, asig_id, prof))
                    lload[prof] += 1
                    _etapa_add(prof, curso_id)
                lunassig.extend([(curso_id, asig_id)] * (count - n))
            else:
                remaining = count
                for prof in sort_eligible(_avail(eligible, curso_id), curso_id, asig_id, lload):
                    if remaining <= 0:
                        break
                    etapa = curso_etapa.get(curso_id)
                    etapa_restante = (rp_horas_etapa.get(prof, {}).get(etapa, 9999)
                                      - letapa[prof].get(etapa, 0)) if etapa else 9999
                    cap = min(prof_max.get(prof, 25) - lload[prof], etapa_restante)
                    n = int(min(remaining, cap))
                    for _ in range(n):
                        lnormal.append((curso_id, asig_id, prof))
                        lload[prof] += 1
                        _etapa_add(prof, curso_id)
                    remaining -= n
                lunassig.extend([(curso_id, asig_id)] * remaining)

        for curso_id, asig_id1, asig_id2 in paired_tasks:
            e1 = eligible_for(curso_id, asig_id1)
            e2 = eligible_for(curso_id, asig_id2)
            a1 = sort_eligible(_avail(e1, curso_id), curso_id, asig_id1, lload)
            a2 = sort_eligible(_avail(e2, curso_id), curso_id, asig_id2, lload)
            prof1 = a1[0] if a1 else None
            prof2 = a2[0] if a2 else prof1
            if prof1:
                lpaired.append((curso_id, asig_id1, prof1, asig_id2, prof2))
                lload[prof1] += 1
                _etapa_add(prof1, curso_id)
                if prof2 and prof2 != prof1:
                    lload[prof2] += 1
                    _etapa_add(prof2, curso_id)
            else:
                lunassig.extend([(curso_id, asig_id1), (curso_id, asig_id2)])

        deficit = sum(max(0, prof_max.get(p.id, 0) - lload[p.id]) for p in profesores_activos)
        return lnormal, lpaired, lunassig, lload, letapa, deficit

    # Reintentar Phase 1 hasta que todos los profesores alcancen su objetivo
    best_p1 = _run_fase1()
    best_deficit = best_p1[5]
    P1_MAX = 500
    p1_sin_mejora = 0
    for _ in range(P1_MAX):
        if best_deficit == 0:
            break
        result = _run_fase1()
        if result[5] < best_deficit:
            best_deficit = result[5]
            best_p1 = result
            p1_sin_mejora = 0
        else:
            p1_sin_mejora += 1
            if p1_sin_mejora >= 200:
                break  # sin progreso: problema estructural en los datos

    p1_normal, p1_paired, p1_unassignable, prof_load, prof_etapa_load, _ = best_p1

    # Advertencias para profesores que no alcanzaron su objetivo
    for p in profesores_activos:
        target = prof_max.get(p.id, 0)
        actual = prof_load[p.id]
        if target > 0 and actual < target:
            p1_unassignable.append(('horas_prof', p.id, actual, target))

    # ══════════════════════════════════════════════════════════════════════
    # FASE 2 — Distribución: ¿en qué día y franja?
    # Corre en bucle hasta encontrar solución completa (máx. 500 intentos).
    # Cada intento usa orden aleatorio diferente; guarda el mejor resultado.
    # ══════════════════════════════════════════════════════════════════════

    slots_all = [(dia, franja) for dia in DIAS_SEMANA for franja in franjas_clase]

    # Mapa tutor → cursos (para regla tutor_primera)
    tutor_cursos = defaultdict(set)
    for cname, pids in prof_tutoria.items():
        for cid, cn in curso_nombre.items():
            if cn == cname:
                for pid in pids:
                    tutor_cursos[pid].add(cid)

    # Descontar de p1_normal las tareas ya cubiertas por celdas manuales
    _mc = defaultdict(int)
    for _key in manual_covered:
        _mc[_key] += 1
    _p1_pendiente = []
    for item in p1_normal:
        _key = (item[0], item[1], item[2])
        if _mc[_key] > 0:
            _mc[_key] -= 1
        else:
            _p1_pendiente.append(item)

    fijar_items  = [(c, a, p) for c, a, p in _p1_pendiente if a in reglas_fijar]
    other_items  = [(c, a, p) for c, a, p in _p1_pendiente if a not in reglas_fijar]
    consec_items = [(c, a, p) for c, a, p in other_items if a in reglas_consecutivas]
    normal_items = [(c, a, p) for c, a, p in other_items if a not in reglas_consecutivas]

    def _slots_for(cid, aid, pid, xpid=None):
        """Slots válidos para un item dado el estado actual."""
        mx = reglas_max_dia.get(aid)
        out = []
        for dia, franja in slots_all:
            sl = (dia, franja)
            if sl in c_occ[cid]: continue
            if sl in p_occ[pid]: continue
            if xpid and xpid != pid and sl in p_occ[xpid]: continue
            if pid in rp_excluir_franja.get(sl, ()): continue
            if xpid and xpid in rp_excluir_franja.get(sl, ()): continue
            if mx and d_cnt[(cid, aid, dia)] >= mx: continue
            out.append(sl)
        return out

    def _consec_seqs(cid, aid, pid, n):
        """Secuencias de n franjas consecutivas válidas para un item."""
        out = []
        for dia in DIAS_SEMANA:
            for si in range(len(franjas_clase) - n + 1):
                seq = [(dia, franjas_clase[si + k]) for k in range(n)]
                if any(s in c_occ[cid] or s in p_occ[pid] for s in seq): continue
                if any(pid in rp_excluir_franja.get(s, ()) for s in seq): continue
                mx = reglas_max_dia.get(aid)
                if mx and d_cnt[(cid, aid, dia)] + n > mx: continue
                out.append(seq)
        return out

    # Estado mutable compartido por el backtracking
    p_occ = defaultdict(set)
    c_occ = defaultdict(set)
    d_cnt = defaultdict(int)

    for _pid, _ss in manual_slots_prof.items():
        p_occ[_pid].update(_ss)
    for _cid, _ss in manual_slots_curso.items():
        c_occ[_cid].update(_ss)

    # Colocar fijar_items de forma determinista (no hay elección)
    placed_base = []
    fallos_base = 0
    for cid, aid, pid in fijar_items:
        fd, ff = reglas_fijar[aid]
        sl = (fd, ff)
        if sl not in c_occ[cid] and sl not in p_occ[pid]:
            placed_base.append((cid, aid, pid, fd, ff, None, None))
            p_occ[pid].add(sl); c_occ[cid].add(sl); d_cnt[(cid, aid, fd)] += 1
        else:
            fallos_base += 1

    # Construir lista de variables para backtracking
    # ('n', cid, aid, pid) | ('c', cid, aid, pid, n) | ('p', cid, a1, p1, a2, p2)
    bt_vars = []
    cg = defaultdict(int)
    for c, a, p in consec_items:
        cg[(c, a, p)] += 1
    for (c, a, p), n in cg.items():
        bt_vars.append(('c', c, a, p, n))
    for c, a, p in normal_items:
        bt_vars.append(('n', c, a, p))
    for c, a1, p1, a2, p2 in p1_paired:
        bt_vars.append(('p', c, a1, p1, a2, p2))

    random.shuffle(bt_vars)

    # ── FASE 2: Greedy con reinicios ─────────────────────────────────────
    # El backtracking es demasiado lento para 200+ variables.
    # Usamos greedy ordenado por slack con muchos reinicios aleatorios.
    # Slack = slots_disponibles - items_pendientes_del_profesor:
    #   Jaione (21 items, 25 slots) → slack=4 → se coloca ANTES
    #   Profesor con 3 items, 25 slots → slack=22 → se coloca DESPUÉS

    import time as _time
    _bt_start = _time.time()
    BT_TIMEOUT = 270

    # Pre-calcular cuántos items hay por (prof, class) combinación
    items_por_prof = defaultdict(int)
    items_por_clase = defaultdict(int)
    for item in bt_vars:
        if item[0] == 'n':
            items_por_prof[item[3]] += 1
            items_por_clase[item[1]] += 1
        elif item[0] == 'p':
            items_por_prof[item[3]] += 1
            items_por_prof[item[5]] += 1
            items_por_clase[item[1]] += 1

    def _initial_slack(item):
        """Slack combinado prof+clase: menor = más urgente."""
        if item[0] == 'n':
            pid, cid = item[3], item[1]
            # Slots válidos para AMBOS prof y clase libres
            avail = sum(1 for d, f in slots_all
                        if (d, f) not in p_occ[pid] and (d, f) not in c_occ[cid])
            # Usar el máximo pendiente entre prof y clase (más restrictivo)
            pending = max(items_por_prof[pid], items_por_clase[cid])
            return avail - pending
        elif item[0] == 'c':
            pid = item[3]
            n = item[4]
            avail = len(_consec_seqs(item[1], item[2], item[3], n))
            return avail - 1
        else:  # paired
            p1, p2 = item[3], item[5]
            avail = sum(1 for d, f in slots_all
                        if (d, f) not in p_occ[p1] and (d, f) not in p_occ[p2]
                        and (d, f) not in c_occ[item[1]])
            return avail - 1

    # Ordenar bt_vars por slack ascendente (más apretado primero)
    bt_vars_sorted = sorted(bt_vars, key=_initial_slack)

    best_greedy = {'fallos': len(bt_vars), 'placed': []}

    def _greedy_pass(order):
        """Un pase greedy sobre los items en el orden dado."""
        local_p_occ = {pid: set(s) for pid, s in p_occ.items()}
        local_c_occ = {cid: set(s) for cid, s in c_occ.items()}
        local_d_cnt = defaultdict(int, d_cnt)
        placed_local = []
        fallos_local = 0

        for item in order:
            if item[0] == 'n':
                _, cid, aid, pid = item
                mx = reglas_max_dia.get(aid)
                slots = [(d, f) for d, f in slots_all
                         if (d, f) not in local_c_occ.get(cid, set())
                         and (d, f) not in local_p_occ.get(pid, set())
                         and pid not in rp_excluir_franja.get((d, f), ())
                         and (not mx or local_d_cnt[(cid, aid, d)] < mx)]
                if not slots:
                    fallos_local += 1
                    continue
                # LCV: preferir el día donde el profesor tenga menos clases ya
                # asignadas (esparcir la carga del profesor por la semana)
                p_day = defaultdict(int)
                for d2, _ in local_p_occ.get(pid, set()):
                    p_day[d2] += 1
                c_day = defaultdict(int)
                for d2, _ in local_c_occ.get(cid, set()):
                    c_day[d2] += 1
                slots.sort(key=lambda s: p_day[s[0]] + 0.5 * c_day[s[0]])
                # Entre los slots con mejor score, elegir al azar
                best_score = p_day[slots[0][0]] + 0.5 * c_day[slots[0][0]]
                candidates = [s for s in slots if p_day[s[0]] + 0.5 * c_day[s[0]] <= best_score + 0.5]
                dia, franja = random.choice(candidates)
                local_p_occ.setdefault(pid, set()).add((dia, franja))
                local_c_occ.setdefault(cid, set()).add((dia, franja))
                local_d_cnt[(cid, aid, dia)] += 1
                placed_local.append((cid, aid, pid, dia, franja, None, None))

            elif item[0] == 'c':
                _, cid, aid, pid, n = item
                seqs = _consec_seqs(cid, aid, pid, n)
                # Filtra con el estado local
                valid_seqs = [s for s in seqs
                              if all((d, f) not in local_c_occ.get(cid, set())
                                     and (d, f) not in local_p_occ.get(pid, set())
                                     for d, f in s)]
                if not valid_seqs:
                    fallos_local += n
                    continue
                seq = random.choice(valid_seqs)
                for dia, franja in seq:
                    local_p_occ.setdefault(pid, set()).add((dia, franja))
                    local_c_occ.setdefault(cid, set()).add((dia, franja))
                    local_d_cnt[(cid, aid, dia)] += 1
                    placed_local.append((cid, aid, pid, dia, franja, None, None))

            else:  # paired
                _, cid, a1, p1, a2, p2 = item
                slots = [(d, f) for d, f in slots_all
                         if (d, f) not in local_c_occ.get(cid, set())
                         and (d, f) not in local_p_occ.get(p1, set())
                         and (d, f) not in local_p_occ.get(p2, set())
                         and p1 not in rp_excluir_franja.get((d, f), ())
                         and p2 not in rp_excluir_franja.get((d, f), ())]
                if not slots:
                    fallos_local += 2
                    continue
                # Preferir días menos cargados para los dos profesores
                p1_day = defaultdict(int)
                for d2, _ in local_p_occ.get(p1, set()):
                    p1_day[d2] += 1
                p2_day = defaultdict(int)
                for d2, _ in local_p_occ.get(p2, set()):
                    p2_day[d2] += 1
                slots.sort(key=lambda s: p1_day[s[0]] + p2_day[s[0]])
                dia, franja = slots[0]
                local_p_occ.setdefault(p1, set()).add((dia, franja))
                if p2 != p1: local_p_occ.setdefault(p2, set()).add((dia, franja))
                local_c_occ.setdefault(cid, set()).add((dia, franja))
                local_d_cnt[(cid, a1, dia)] += 1
                placed_local.append((cid, a1, p1, dia, franja, a2, p2))

        return placed_local, fallos_local

    def _repair_pass(placed_in, fallos_items_in, order):
        """Reparación local: para cada item fallido, intenta reubicar
        el item que le bloquea en otro slot libre."""
        local_p_occ = {pid: set(s) for pid, s in p_occ.items()}
        local_c_occ = {cid: set(s) for cid, s in c_occ.items()}
        local_d_cnt = defaultdict(int, d_cnt)
        placed = list(placed_in)

        # Reconstruir estado a partir de placed
        placed_slots = {}  # (cid,aid,pid,dia,franja) -> index in placed
        for idx, (cid, aid, pid, dia, franja, _, _) in enumerate(placed):
            local_p_occ.setdefault(pid, set()).add((dia, franja))
            local_c_occ.setdefault(cid, set()).add((dia, franja))
            local_d_cnt[(cid, aid, dia)] += 1
            placed_slots[(cid, aid, pid, dia, franja)] = idx

        repaired = 0
        for item in order:
            if item[0] != 'n':
                continue
            _, cid, aid, pid = item
            # ¿Ya está colocado?
            already = any(c == cid and a == aid and p == pid
                          for c, a, p, *_ in placed)
            if already:
                continue
            # Intentar colocarlo moviendo bloqueadores
            mx = reglas_max_dia.get(aid)
            for dia, franja in slots_all:
                sl = (dia, franja)
                if pid in rp_excluir_franja.get(sl, ()):
                    continue
                if mx and local_d_cnt[(cid, aid, dia)] >= mx:
                    continue
                # Encontrar bloqueadores en este slot
                blockers = []
                if sl in local_p_occ.get(pid, set()):
                    # Buscar qué item ocupa este slot del profesor
                    blocker = next((it for it in placed
                                    if it[2] == pid and (it[3], it[4]) == sl), None)
                    if blocker:
                        blockers.append(blocker)
                if sl in local_c_occ.get(cid, set()):
                    blocker = next((it for it in placed
                                    if it[0] == cid and (it[3], it[4]) == sl), None)
                    if blocker and blocker not in blockers:
                        blockers.append(blocker)

                if len(blockers) != 1:
                    continue  # Solo reparar cuando hay un único bloqueador

                b = blockers[0]
                bcid, baid, bpid, bdia, bfranja = b[0], b[1], b[2], b[3], b[4]
                # Intentar mover el bloqueador a otro slot
                alt_slots = [s for s in slots_all if s != (bdia, bfranja)
                             and s not in local_p_occ.get(bpid, set())
                             and s not in local_c_occ.get(bcid, set())
                             and bpid not in rp_excluir_franja.get(s, ())]
                if not alt_slots:
                    continue
                # Elegir mejor slot alternativo (día menos cargado)
                p_day = defaultdict(int)
                for d2, _ in local_p_occ.get(bpid, set()):
                    p_day[d2] += 1
                alt_slots.sort(key=lambda s: p_day[s[0]])
                new_dia, new_franja = alt_slots[0]

                # Mover bloqueador
                placed.remove(b)
                local_p_occ[bpid].discard((bdia, bfranja))
                local_c_occ[bcid].discard((bdia, bfranja))
                local_d_cnt[(bcid, baid, bdia)] -= 1
                local_p_occ.setdefault(bpid, set()).add((new_dia, new_franja))
                local_c_occ.setdefault(bcid, set()).add((new_dia, new_franja))
                local_d_cnt[(bcid, baid, new_dia)] += 1
                placed.append((bcid, baid, bpid, new_dia, new_franja, b[5], b[6]))

                # Colocar el item fallido
                local_p_occ.setdefault(pid, set()).add(sl)
                local_c_occ.setdefault(cid, set()).add(sl)
                local_d_cnt[(cid, aid, dia)] += 1
                placed.append((cid, aid, pid, dia, franja, None, None))
                repaired += 1
                break

        return placed, repaired

    # Reinicios: slack-ordenado + perturbación + reparación local
    iteracion = 0
    while _time.time() - _bt_start < BT_TIMEOUT:
        iteracion += 1
        if iteracion == 1:
            order = bt_vars_sorted[:]
        elif iteracion % 30 == 0:
            # Reinicio completo aleatorio
            order = bt_vars[:]
            random.shuffle(order)
        else:
            order = bt_vars_sorted[:]
            for i in range(len(order) - 1):
                if random.random() < 0.20:
                    j = random.randint(i, min(i + 12, len(order) - 1))
                    order[i], order[j] = order[j], order[i]

        placed_local, fallos_local = _greedy_pass(order)

        # Reparación local si hay fallos
        if fallos_local > 0 and fallos_local <= 20:
            # Calcular items fallidos
            placed_keys_local = set()
            for c, a, p, *_ in placed_local:
                placed_keys_local.add((c, a, p))
            failed_items = [it for it in order
                            if it[0] == 'n' and (it[1], it[2], it[3]) not in placed_keys_local]
            placed_local, n_repaired = _repair_pass(placed_local, failed_items, failed_items)
            fallos_local -= n_repaired

        with _gen_lock:
            _gen_estado['intentos'] = iteracion
            _gen_estado['mejor_fallos'] = best_greedy['fallos']

        if fallos_local < best_greedy['fallos']:
            best_greedy['fallos'] = fallos_local
            best_greedy['placed'] = placed_local[:]
            with _gen_lock:
                _gen_estado['mejor_fallos'] = fallos_local
            if fallos_local == 0:
                break

    best_assignments = placed_base + best_greedy['placed']
    best_fallos = fallos_base + best_greedy['fallos']

    # Escribir el mejor resultado a la base de datos
    for curso_id, asig_id, prof_id, dia, franja, asig2, prof2 in best_assignments:
        ha = HorarioAsignacion(
            curso_id=curso_id, asignatura_id=asig_id,
            profesor_id=prof_id, dia=dia, franja=franja
        )
        if asig2:
            ha.asignatura2_id = asig2
            ha.profesor2_id = prof2 if prof2 else prof_id
        db.session.add(ha)

    db.session.commit()

    # Diagnóstico: identificar qué slots del mejor intento quedaron sin franja
    placed_keys = {(c, a, p) for c, a, p, *_ in best_assignments}
    needed_keys = [(c, a, p) for c, a, p in p1_normal]
    diagnostico = []
    if best_fallos > 0:
        placed_cnt = defaultdict(int)
        for c, a, p, *_ in best_assignments:
            placed_cnt[(c, a, p)] += 1
        needed_cnt = defaultdict(int)
        for c, a, p in p1_normal:
            needed_cnt[(c, a, p)] += 1
        for (c, a, p), n in needed_cnt.items():
            colocados = placed_cnt.get((c, a, p), 0)
            if colocados < n:
                curso_n = curso_nombre.get(c, str(c))
                asig_n = asig_map[a].nombre if a in asig_map else str(a)
                prof_n = next((x.nombre for x in profesores_activos if x.id == p), str(p))
                diagnostico.append(f'{asig_n} en {curso_n} ({prof_n}): {colocados}/{n}')

    n_manuales = len(manual_asigs)
    return p1_unassignable, best_fallos, diagnostico, n_manuales


# ─── CONSTRUCTOR DE HORARIOS ───

@app.route('/horarios-construccion')
@login_required
def horarios_construccion():
    if not current_user.es_admin:
        flash('Solo administradores.', 'danger')
        return redirect(url_for('dashboard'))
    tab = request.args.get('tab', 'asignaturas')
    curso_id = request.args.get('curso_id', type=int)
    vista = request.args.get('vista', 'clase')  # 'clase' o 'profesor'
    prof_id = request.args.get('prof_id', type=int)
    if request.args.get('ok'):
        flash('Horario generado correctamente.', 'success')
    elif request.args.get('aviso'):
        flash('Horario generado con advertencias — ' + request.args.get('aviso'), 'warning')

    asignaturas = Asignatura.query.order_by(Asignatura.nombre).all()
    # Pre-agrupar por etapa en Python para evitar problemas con None en Jinja2 groupby
    _etapa_order = {'Haur Hezkuntza': 0, 'Lehen Hezkuntza': 1, '1-2 años': 2, 'Otras': 3}
    _grupos = {}
    for a in asignaturas:
        key = a.etapa or ''
        _grupos.setdefault(key, []).append(a)
    asignaturas_por_etapa = sorted(_grupos.items(), key=lambda x: _etapa_order.get(x[0], 99))
    cursos = Curso.query.order_by(Curso.orden, Curso.nombre).all()
    profesores_lista = Profesor.query.filter_by(activo=True, de_baja=False, es_admin=False).order_by(Profesor.nombre).all()

    curso_sel = None
    horario_grid = {}
    prof_sel = None
    horario_prof_grid = {}
    comp_grid = {}
    prof_grupos = []
    if tab == 'horario':
        if vista == 'profesor' and profesores_lista:
            prof_sel = Profesor.query.get(prof_id) if prof_id else profesores_lista[0]
            if prof_sel:
                for a in HorarioAsignacion.query.filter_by(profesor_id=prof_sel.id).all():
                    horario_prof_grid[(a.dia, a.franja)] = a
                for s in SlotComplementaria.query.filter_by(profesor_id=prof_sel.id).all():
                    comp_grid[(s.dia, s.franja)] = s
                prof_grupos = (GrupoTrabajo.query
                               .join(ProfesorGrupo, ProfesorGrupo.grupo_id == GrupoTrabajo.id)
                               .filter(ProfesorGrupo.profesor_id == prof_sel.id)
                               .order_by(GrupoTrabajo.nombre).all())
        elif cursos:
            curso_sel = Curso.query.get(curso_id) if curso_id else cursos[0]
            if curso_sel:
                for a in HorarioAsignacion.query.filter_by(curso_id=curso_sel.id).all():
                    horario_grid[(a.dia, a.franja)] = a

    prof_asignaturas = {
        p.id: [pa.asignatura_id for pa in ProfesorAsignatura.query.filter_by(profesor_id=p.id).all()]
        for p in profesores_lista
    }
    prof_horas_asig = {
        p.id: HorarioAsignacion.query.filter_by(profesor_id=p.id).count()
        for p in profesores_lista
    }
    prof_etapas_list = {
        p.id: parse_etapas(p.etapa)
        for p in profesores_lista
    }
    prof_especialidades = {
        p.id: ProfesorEspecialidad.query.filter_by(profesor_id=p.id).order_by(ProfesorEspecialidad.nombre).all()
        for p in profesores_lista
    }

    reglas = ReglaHorario.query.all()
    franjas_clase = [f for f in FRANJAS if f != 'Patio']

    # Asignaturas sin ningún profesor asignado (para aviso en tab Horario)
    profs_por_asig = defaultdict(set)
    for pa in ProfesorAsignatura.query.all():
        profs_por_asig[pa.asignatura_id].add(pa.profesor_id)
    cursos_map_local = {c.id: c for c in cursos}
    asig_map_local = {a.id: a for a in asignaturas}
    sin_profesor = []
    for req in CursoAsignatura.query.all():
        if not profs_por_asig.get(req.asignatura_id):
            curso_n = cursos_map_local.get(req.curso_id)
            asig_n = asig_map_local.get(req.asignatura_id)
            if curso_n and asig_n:
                sin_profesor.append(f'{asig_n.nombre} ({curso_n.nombre})')
    sin_profesor.sort()

    return render_template('horarios_construccion.html',
        tab=tab, vista=vista, asignaturas=asignaturas, asignaturas_por_etapa=asignaturas_por_etapa,
        cursos=cursos, profesores=profesores_lista,
        curso_sel=curso_sel, horario_grid=horario_grid,
        prof_sel=prof_sel, horario_prof_grid=horario_prof_grid,
        prof_asignaturas=prof_asignaturas,
        prof_horas_asig=prof_horas_asig, prof_especialidades=prof_especialidades,
        prof_etapas_list=prof_etapas_list,
        etapas=ETAPAS, dias=DIAS_SEMANA, franjas=FRANJAS, franjas_clase=franjas_clase,
        sin_profesor=sin_profesor, comp_grid=comp_grid,
        prof_grupos=prof_grupos,
        reglas=reglas)


@app.route('/horarios-construccion/asignatura/nueva', methods=['POST'])
@login_required
def nueva_asignatura_hc():
    if not current_user.es_admin:
        return redirect(url_for('dashboard'))
    nombre = request.form.get('nombre', '').strip()
    color = request.form.get('color', '#0d6efd').strip()
    etapa = request.form.get('etapa', '').strip() or None
    if nombre and not Asignatura.query.filter_by(nombre=nombre).first():
        db.session.add(Asignatura(nombre=nombre, color=color, etapa=etapa))
        db.session.commit()
        flash(f'Asignatura "{nombre}" añadida.', 'success')
    elif nombre:
        flash('Ya existe una asignatura con ese nombre.', 'warning')
    return redirect(url_for('horarios_construccion', tab='asignaturas'))


@app.route('/horarios-construccion/asignatura/<int:id>/eliminar', methods=['POST'])
@login_required
def eliminar_asignatura_hc(id):
    if not current_user.es_admin:
        return redirect(url_for('dashboard'))
    a = Asignatura.query.get_or_404(id)
    CursoAsignatura.query.filter_by(asignatura_id=id).delete()
    ProfesorAsignatura.query.filter_by(asignatura_id=id).delete()
    HorarioAsignacion.query.filter_by(asignatura_id=id).delete()
    db.session.delete(a)
    db.session.commit()
    flash('Asignatura eliminada.', 'success')
    return redirect(url_for('horarios_construccion', tab='asignaturas'))


@app.route('/horarios-construccion/requisito', methods=['POST'])
@login_required
def guardar_requisito_hc():
    if not current_user.es_admin:
        return redirect(url_for('dashboard'))
    curso_id = int(request.form.get('curso_id'))
    asignatura_id = int(request.form.get('asignatura_id'))
    horas = float(request.form.get('horas', 1))
    accion = request.form.get('accion', 'guardar')
    if accion == 'eliminar':
        CursoAsignatura.query.filter_by(
            curso_id=curso_id, asignatura_id=asignatura_id).delete()
        db.session.commit()
    else:
        req = CursoAsignatura.query.filter_by(
            curso_id=curso_id, asignatura_id=asignatura_id).first()
        if req:
            req.horas_semanales = horas
        else:
            db.session.add(CursoAsignatura(
                curso_id=curso_id, asignatura_id=asignatura_id, horas_semanales=horas))
        db.session.commit()
    return redirect(url_for('horarios_construccion', tab='asignaturas'))


@app.route('/horarios-construccion/asignatura/<int:asig_id>/requisitos', methods=['POST'])
@login_required
def guardar_requisitos_asignatura_hc(asig_id):
    if not current_user.es_admin:
        return redirect(url_for('dashboard'))
    asig = Asignatura.query.get_or_404(asig_id)
    curso_ids_marcados = [int(x) for x in request.form.getlist('curso_ids')]
    CursoAsignatura.query.filter_by(asignatura_id=asig_id).delete()
    for curso_id in curso_ids_marcados:
        horas = float(request.form.get(f'horas_{curso_id}', 1))
        if horas > 0:
            db.session.add(CursoAsignatura(
                curso_id=curso_id, asignatura_id=asig_id, horas_semanales=horas))
    db.session.commit()
    flash(f'Requisitos de {asig.nombre} guardados.', 'success')
    return redirect(url_for('horarios_construccion', tab='asignaturas'))


@app.route('/horarios-construccion/profesor/<int:prof_id>/config', methods=['POST'])
@login_required
def config_profesor_hc(prof_id):
    if not current_user.es_admin:
        return redirect(url_for('dashboard'))
    p = Profesor.query.get_or_404(prof_id)
    p.horas_max_semanales = int(request.form.get('horas_max', 25))
    asigs_sel = request.form.getlist('asignaturas')
    ProfesorAsignatura.query.filter_by(profesor_id=prof_id).delete()
    for asig_id in asigs_sel:
        db.session.add(ProfesorAsignatura(
            profesor_id=prof_id, asignatura_id=int(asig_id)))
    db.session.commit()
    flash(f'Configuración de {p.nombre} guardada.', 'success')
    return redirect(url_for('horarios_construccion', tab='profesores'))


@app.route('/horarios-construccion/profesor/<int:prof_id>/especialidades', methods=['POST'])
@login_required
def guardar_especialidades_hc(prof_id):
    if not current_user.es_admin:
        return redirect(url_for('dashboard'))
    p = Profesor.query.get_or_404(prof_id)
    ProfesorEspecialidad.query.filter_by(profesor_id=prof_id).delete()
    nombres = request.form.getlist('esp_nombre')
    horas_list = request.form.getlist('esp_horas')
    for nombre, horas in zip(nombres, horas_list):
        nombre = nombre.strip()
        if nombre:
            try:
                h = max(0, float(horas))
            except ValueError:
                h = 0
            db.session.add(ProfesorEspecialidad(
                profesor_id=prof_id, nombre=nombre, horas_semanales=h))
    db.session.commit()
    flash(f'Especialidades de {p.nombre} guardadas.', 'success')
    return redirect(url_for('horarios_construccion', tab='profesores'))


def _worker_generar():
    with app.app_context():
        try:
            sin_asignar_p1, fallos_p2, diagnostico, n_manuales = generar_horario_automatico()
            _, grupos_sin_hueco = _asignar_grupos_trabajo()
            avisos = []
            if n_manuales > 0:
                avisos.append(f'ℹ️ {n_manuales} celda(s) manual(es) preservada(s)')
            slots_sin_asignar = [x for x in sin_asignar_p1 if x[0] != 'horas_prof']
            horas_prof_cortas = [x for x in sin_asignar_p1 if x[0] == 'horas_prof']
            if slots_sin_asignar:
                nombres = []
                for curso_id, asig_id in slots_sin_asignar[:5]:
                    c = Curso.query.get(curso_id)
                    a = Asignatura.query.get(asig_id)
                    if c and a:
                        nombres.append(f'{a.nombre} ({c.nombre})')
                if nombres:
                    avisos.append(f'Sin asignar: {", ".join(nombres)}')
            if horas_prof_cortas:
                msgs = []
                for _, pid, actual, target in horas_prof_cortas[:5]:
                    p = Profesor.query.get(pid)
                    if p:
                        msgs.append(f'{p.nombre} ({actual}/{target}h)')
                if msgs:
                    avisos.append(f'Horas insuficientes: {", ".join(msgs)}')
            if diagnostico:
                avisos.append(f'Conflicto de horario (sin franja libre): {", ".join(diagnostico[:5])}')
            if grupos_sin_hueco:
                avisos.append(f'Sin hueco común: {", ".join(grupos_sin_hueco)}')
            with _gen_lock:
                _gen_estado['avisos'] = avisos
                _gen_estado['listo'] = True
                _gen_estado['activo'] = False
        except Exception as e:
            with _gen_lock:
                _gen_estado['error'] = str(e)
                _gen_estado['activo'] = False
                _gen_estado['listo'] = True


@app.route('/horarios-construccion/generar', methods=['POST'])
@login_required
def generar_horario_hc():
    if not current_user.es_admin:
        return redirect(url_for('dashboard'))
    with _gen_lock:
        if _gen_estado['activo']:
            flash('Ya hay una generación en curso, espera.', 'warning')
            return redirect(url_for('horario_generando'))
        _gen_estado.update({'activo': True, 'intentos': 0, 'mejor_fallos': None,
                            'error': None, 'avisos': [], 'listo': False})
    t = threading.Thread(target=_worker_generar, daemon=True)
    t.start()
    return redirect(url_for('horario_generando'))


@app.route('/horarios-construccion/generando')
@login_required
def horario_generando():
    if not current_user.es_admin:
        return redirect(url_for('dashboard'))
    return render_template('horario_generando.html')


@app.route('/horarios-construccion/estado-generacion')
@login_required
def estado_generacion():
    with _gen_lock:
        estado = dict(_gen_estado)
    return jsonify(estado)


@app.route('/horarios-construccion/diagnostico')
@login_required
def diagnostico_horario():
    """Simula la Fase 1 con reglas y devuelve un diagnóstico detallado."""
    if not current_user.es_admin:
        return jsonify({'error': 'Solo administradores'}), 403

    profesores_activos = Profesor.query.filter_by(activo=True, de_baja=False, es_admin=False).all()
    cursos_list = Curso.query.order_by(Curso.orden, Curso.nombre).all()
    cursos_activos = [c for c in cursos_list if not c.aula_cerrada]

    prof_por_asig = defaultdict(list)
    for pa in ProfesorAsignatura.query.all():
        prof_por_asig[pa.asignatura_id].append(pa.profesor_id)

    prof_etapas_d = {p.id: parse_etapas(p.etapa) for p in profesores_activos}
    asig_map_loc = {a.id: a for a in Asignatura.query.all()}
    curso_etapa_d = {c.id: c.etapa for c in cursos_list}
    curso_nombre_d = {c.id: c.nombre for c in cursos_list}

    prof_max_d = {
        p.id: float(p.horas_lectivas) if (p.horas_lectivas or 0) > 0 else (p.horas_max_semanales or 25)
        for p in profesores_activos
    }

    # ── Cargar reglas de tutor (igual que en generar_horario_automatico) ──
    reglas_tutor_etapa_dura = defaultdict(set)   # etapa -> set(asig_id)
    reglas_tutor_etapa_blanda = defaultdict(set)
    rp_fijar_curso_asig = defaultdict(set)
    for r in ReglaHorario.query.all():
        if r.tipo == 'tutor_clase_etapa' and r.etapa and r.asignatura_id:
            if r.dureza == 'dura':
                reglas_tutor_etapa_dura[r.etapa].add(r.asignatura_id)
            else:
                reglas_tutor_etapa_blanda[r.etapa].add(r.asignatura_id)
        elif r.tipo == 'prof_fijar_curso_asignatura' and r.curso_id_regla and r.asignatura_id and r.profesor_id:
            rp_fijar_curso_asig[(r.curso_id_regla, r.asignatura_id)].add(r.profesor_id)

    prof_tutoria_d = defaultdict(list)
    for p in profesores_activos:
        if p.aula_tutoria:
            prof_tutoria_d[p.aula_tutoria].append(p.id)

    def etapa_compat(prof_id, asig_id):
        asig = asig_map_loc.get(asig_id)
        if not asig or not asig.etapa:
            return True
        etapas_p = prof_etapas_d.get(prof_id, [])
        if not etapas_p:
            return True
        return asig.etapa in etapas_p

    def elegibles_para(curso_id, asig_id):
        """Replica eligible_for() del generador, incluyendo reglas de tutor."""
        fij_ca = rp_fijar_curso_asig.get((curso_id, asig_id), set())
        if fij_ca:
            return list(fij_ca)
        profs = {pid for pid in prof_por_asig.get(asig_id, [])
                 if etapa_compat(pid, asig_id)}
        etapa = curso_etapa_d.get(curso_id)
        if etapa and asig_id in reglas_tutor_etapa_dura.get(etapa, set()):
            tutores = set(prof_tutoria_d.get(curso_nombre_d.get(curso_id, ''), []))
            profs &= tutores
        return list(profs)

    # ── Construir tareas ──────────────────────────────────────────────────
    tareas = []
    for ca in CursoAsignatura.query.all():
        curso = next((c for c in cursos_activos if c.id == ca.curso_id), None)
        if not curso:
            continue
        h = ca.horas_semanales or 0
        if h > 0:
            tareas.append((ca.curso_id, ca.asignatura_id, h))

    # ── Simulación greedy de Fase 1 con reglas ────────────────────────────
    load_sim = defaultdict(float)
    asig_sim = []
    sin_asignar = []

    tareas_ord = sorted(tareas, key=lambda t: len(elegibles_para(t[0], t[1])))

    for cid, aid, horas in tareas_ord:
        elegibles = elegibles_para(cid, aid)
        if not elegibles:
            razon = 'Sin profesor asignado a esta asignatura'
            # Detectar si es por regla de tutor sin tutor configurado
            etapa = curso_etapa_d.get(cid)
            if etapa and aid in reglas_tutor_etapa_dura.get(etapa, set()):
                cn = curso_nombre_d.get(cid, '')
                tutores_curso = prof_tutoria_d.get(cn, [])
                if not tutores_curso:
                    razon = f'Regla de tutor dura activa pero {cn} no tiene tutor configurado'
                else:
                    razon = f'Regla de tutor dura: tutor(es) de {cn} no tienen esta asignatura'
            sin_asignar.append((cid, aid, horas, razon))
            continue
        elegibles_ord = sorted(elegibles, key=lambda p: load_sim[p] - prof_max_d.get(p, 25))
        asignado = 0.0
        for pid in elegibles_ord:
            restante = prof_max_d.get(pid, 25) - load_sim[pid]
            if restante <= 0:
                continue
            tomar = min(horas - asignado, restante)
            load_sim[pid] += tomar
            asig_sim.append((cid, aid, pid, tomar))
            asignado += tomar
            if asignado >= horas:
                break
        if asignado < horas - 0.01:
            sin_asignar.append((cid, aid, horas - asignado,
                                f'Profesores al límite (asignado {asignado:.1f}/{horas:.1f}h)'))

    # ── Resultado ─────────────────────────────────────────────────────────
    problemas = []
    avisos = []

    if sin_asignar:
        for cid, aid, faltan, razon in sin_asignar:
            cn = curso_nombre_d.get(cid, f'Curso {cid}')
            an = asig_map_loc.get(aid)
            an = an.nombre if an else f'Asig {aid}'
            problemas.append({'tipo': 'sin_asignar',
                               'texto': f'{an} en {cn}: faltan {faltan:.1f}h — {razon}'})

    for p in profesores_activos:
        if prof_max_d.get(p.id, 0) > 0 and load_sim[p.id] == 0:
            total_asigs = ProfesorAsignatura.query.filter_by(profesor_id=p.id).count()
            msg = f'{p.nombre} (máx {prof_max_d[p.id]}h) — ' + (
                'no tiene asignaturas asignadas' if total_asigs == 0
                else 'sus asignaturas las cubren otros profesores (o regla de tutor)')
            (problemas if total_asigs == 0 else avisos).append({'tipo': 'prof_vacio', 'texto': msg})

    resumen_profs = []
    for p in sorted(profesores_activos, key=lambda x: x.nombre):
        asignadas = round(load_sim[p.id], 1)
        maximo = prof_max_d.get(p.id, 25)
        n_asigs = ProfesorAsignatura.query.filter_by(profesor_id=p.id).count()
        resumen_profs.append({
            'nombre': p.nombre,
            'asignadas': asignadas,
            'maximo': maximo,
            'n_asignaturas': n_asigs,
            'ok': asignadas >= maximo * 0.9
        })

    etapa_necesarias = defaultdict(float)
    etapa_asignadas_d = defaultdict(float)
    for cid, aid, h in tareas:
        etapa_necesarias[curso_etapa_d.get(cid, 'Sin etapa') or 'Sin etapa'] += h
    for cid, aid, pid, h in asig_sim:
        etapa_asignadas_d[curso_etapa_d.get(cid, 'Sin etapa') or 'Sin etapa'] += h

    resumen_etapas = []
    for etapa in sorted(etapa_necesarias):
        nec = round(etapa_necesarias[etapa], 1)
        asig = round(etapa_asignadas_d.get(etapa, 0), 1)
        resumen_etapas.append({'etapa': etapa, 'necesarias': nec, 'asignadas': asig,
                                'ok': asig >= nec - 0.1})

    return jsonify({
        'problemas': problemas,
        'avisos': avisos,
        'resumen_etapas': resumen_etapas,
        'resumen_profs': resumen_profs
    })


@app.route('/horarios-construccion/asignar-profesor', methods=['POST'])
@login_required
def asignar_franja_prof_hc():
    if not current_user.es_admin:
        return redirect(url_for('dashboard'))
    profesor_id = int(request.form.get('profesor_id'))
    dia = request.form.get('dia')
    franja = request.form.get('franja')
    curso_id = request.form.get('curso_id', '') or None
    asignatura_id = request.form.get('asignatura_id', '') or None
    asignacion_id = request.form.get('asignacion_id', '') or None

    if asignacion_id:
        ha = HorarioAsignacion.query.get(int(asignacion_id))
    else:
        ha = HorarioAsignacion.query.filter_by(profesor_id=profesor_id, dia=dia, franja=franja).first()

    if not curso_id or not asignatura_id:
        if ha:
            db.session.delete(ha)
            db.session.commit()
    else:
        if ha:
            ha.curso_id = int(curso_id)
            ha.asignatura_id = int(asignatura_id)
            ha.es_manual = True
        else:
            db.session.add(HorarioAsignacion(
                profesor_id=profesor_id, dia=dia, franja=franja,
                curso_id=int(curso_id), asignatura_id=int(asignatura_id),
                es_manual=True
            ))
        db.session.commit()
    return redirect(url_for('horarios_construccion', tab='horario', vista='profesor', prof_id=profesor_id))


@app.route('/horarios-construccion/asignar-slot-comp', methods=['POST'])
@login_required
def asignar_slot_comp_hc():
    """Asigna/edita/borra un SlotComplementaria manualmente (grupo o libre)."""
    if not current_user.es_admin:
        return redirect(url_for('dashboard'))
    profesor_id = int(request.form.get('profesor_id'))
    dia = request.form.get('dia')
    franja = request.form.get('franja')
    tipo = request.form.get('tipo_slot', '')  # 'grupo' | 'libre' | ''
    grupo_id = request.form.get('grupo_id', '') or None
    slot_id = request.form.get('slot_id', '') or None

    existing = (SlotComplementaria.query.get(int(slot_id)) if slot_id
                else SlotComplementaria.query.filter_by(profesor_id=profesor_id, dia=dia, franja=franja).first())

    if not tipo:
        if existing:
            db.session.delete(existing)
            db.session.commit()
    else:
        if existing:
            existing.tipo = tipo
            existing.grupo_id = int(grupo_id) if grupo_id else None
            existing.tipo2 = None
            existing.es_manual = True
        else:
            db.session.add(SlotComplementaria(
                profesor_id=profesor_id, dia=dia, franja=franja,
                tipo=tipo,
                grupo_id=int(grupo_id) if grupo_id else None,
                es_manual=True
            ))
        db.session.commit()
    return redirect(url_for('horarios_construccion', tab='horario', vista='profesor', prof_id=profesor_id))


@app.route('/horarios-construccion/asignar', methods=['POST'])
@login_required
def asignar_franja_hc():
    if not current_user.es_admin:
        return redirect(url_for('dashboard'))
    curso_id = int(request.form.get('curso_id'))
    dia = request.form.get('dia')
    franja = request.form.get('franja')
    asignatura_id = request.form.get('asignatura_id', '')
    profesor_id = request.form.get('profesor_id', '')

    existing = HorarioAsignacion.query.filter_by(
        curso_id=curso_id, dia=dia, franja=franja).first()

    if not asignatura_id or not profesor_id:
        if existing:
            db.session.delete(existing)
            db.session.commit()
    else:
        if existing:
            existing.asignatura_id = int(asignatura_id)
            existing.profesor_id = int(profesor_id)
            existing.es_manual = True
        else:
            db.session.add(HorarioAsignacion(
                curso_id=curso_id, dia=dia, franja=franja,
                asignatura_id=int(asignatura_id),
                profesor_id=int(profesor_id),
                es_manual=True
            ))
        db.session.commit()
    return redirect(url_for('horarios_construccion', tab='horario', curso_id=curso_id))


@app.route('/horarios-construccion/limpiar', methods=['POST'])
@login_required
def limpiar_horario_hc():
    if not current_user.es_admin:
        return redirect(url_for('dashboard'))
    HorarioAsignacion.query.delete()
    SlotComplementaria.query.delete()
    db.session.commit()
    flash('Horario borrado completamente.', 'success')
    return redirect(url_for('horarios_construccion', tab='horario'))


@app.route('/horarios-construccion/limpiar-automaticos', methods=['POST'])
@login_required
def limpiar_automaticos_hc():
    if not current_user.es_admin:
        return redirect(url_for('dashboard'))
    HorarioAsignacion.query.filter_by(es_manual=False).delete()
    try:
        SlotComplementaria.query.filter_by(es_manual=False).delete()
    except Exception:
        pass
    db.session.commit()
    flash('Slots automáticos borrados. Los bloqueados (🔒) se han conservado.', 'success')
    return redirect(url_for('horarios_construccion', tab='horario'))


def _asignar_grupos_trabajo():
    """Asigna slots de grupos de trabajo y complementarias individuales."""
    import random, math
    SlotComplementaria.query.filter_by(es_manual=False).delete()

    franjas_clase = [f for f in FRANJAS if f != 'Patio']
    slots_all = [(dia, franja) for dia in DIAS_SEMANA for franja in franjas_clase]

    # Slots ya ocupados por clases para cada profesor
    prof_ocupado = defaultdict(set)
    for asig in HorarioAsignacion.query.all():
        prof_ocupado[asig.profesor_id].add((asig.dia, asig.franja))
        if asig.profesor2_id:
            prof_ocupado[asig.profesor2_id].add((asig.dia, asig.franja))

    # Slots manuales de complementaria ya colocados (no se regeneran ni se solapan)
    for sc in SlotComplementaria.query.filter_by(es_manual=True).all():
        prof_ocupado[sc.profesor_id].add((sc.dia, sc.franja))

    # Slots de grupos ya asignados a cada profesor (para descontar de complementarias)
    prof_slots_grupo = defaultdict(int)

    grupos = GrupoTrabajo.query.order_by(GrupoTrabajo.id).all()
    grupos_sin_hueco = []
    for grupo in grupos:
        miembros = ProfesorGrupo.query.filter_by(grupo_id=grupo.id).all()
        if not miembros:
            continue

        # Regla dura: todos los miembros coinciden en los mismos slots.
        # Se buscan tantos slots compartidos como el máximo de horas (redondeado arriba).
        max_horas = math.ceil(max(m.horas_semanales for m in miembros))
        slots_s = slots_all[:]
        random.shuffle(slots_s)
        prof_ids = [m.profesor_id for m in miembros]
        miembro_map = {m.profesor_id: m for m in miembros}

        shared_slots = []
        for dia, franja in slots_s:
            if len(shared_slots) >= max_horas:
                break
            slot = (dia, franja)
            if all(slot not in prof_ocupado[pid] for pid in prof_ids):
                shared_slots.append(slot)
                for pid in prof_ids:
                    prof_ocupado[pid].add(slot)

        if len(shared_slots) < max_horas:
            grupos_sin_hueco.append(f'{grupo.nombre} ({len(shared_slots)}/{max_horas})')

        # Crear slots por miembro según sus horas individuales
        for pid in prof_ids:
            m = miembro_map[pid]
            n_slots = math.ceil(m.horas_semanales)
            tiene_fraccion = (m.horas_semanales % 1) >= 0.5
            for i, (dia, franja) in enumerate(shared_slots[:n_slots]):
                es_ultimo = (i == n_slots - 1)
                tipo2 = 'media' if (es_ultimo and tiene_fraccion) else None
                sc = SlotComplementaria(
                    profesor_id=pid, dia=dia, franja=franja,
                    tipo='grupo', grupo_id=grupo.id, tipo2=tipo2
                )
                db.session.add(sc)
                prof_slots_grupo[pid] += 1

    # Flush para que los slots de grupo estén disponibles
    db.session.flush()

    # ── Parear fracciones de media hora ──────────────────────────────────────
    # Regla: ningún slot puede quedar ocupado solo por media hora.
    # Si un slot de grupo es tipo2='media' (½ grupo) necesita pareja de ½ libre.
    # Si hay fracción de horas_libres también, se combinan en un solo slot.

    # 1. Recoger qué profes tienen fracción de grupo (slot tipo2='media')
    prof_grupo_media = {}   # pid -> SlotComplementaria (el slot 'media')
    for sc in SlotComplementaria.query.filter_by(tipo='grupo', tipo2='media').all():
        if not sc.es_manual:
            prof_grupo_media[sc.profesor_id] = sc

    # 2. Recoger fracción de horas libres por profesor
    prof_libre_fraccion = set()
    for prof in Profesor.query.filter_by(activo=True, de_baja=False, es_admin=False).all():
        horas = prof.horas_libres or 0
        fraccion = horas - int(horas)
        if fraccion >= 0.5:
            prof_libre_fraccion.add(prof.id)

    # 3. Parear: fracción grupo + fracción libre → un solo slot tipo2='libre'
    prof_libre_ya_pareado = set()
    for pid in list(prof_grupo_media.keys()):
        if pid in prof_libre_fraccion:
            prof_grupo_media[pid].tipo2 = 'libre'   # ½ grupo + ½ libre
            prof_libre_ya_pareado.add(pid)
            del prof_grupo_media[pid]

    # 4. Fracción grupo sin pareja libre → convertir a 'libre' igualmente
    #    (la mitad vacía se trata como complementaria libre)
    for pid, sc in prof_grupo_media.items():
        sc.tipo2 = 'libre'

    # 5. Fracción libre sin slot de grupo → buscar cualquier slot de grupo o crear ½ libre
    for pid in prof_libre_fraccion:
        if pid in prof_libre_ya_pareado:
            continue
        # Intentar pegar la ½ libre a un slot de grupo completo (tipo2=None)
        candidato = None
        for gs in SlotComplementaria.query.filter_by(profesor_id=pid, tipo='grupo').all():
            if not gs.es_manual and gs.tipo2 is None:
                candidato = gs
                break
        if candidato:
            candidato.tipo2 = 'libre'
        else:
            # Crear un slot ½ libre independiente si hay hueco
            for dia, franja in slots_all:
                sl = (dia, franja)
                if sl not in prof_ocupado[pid]:
                    db.session.add(SlotComplementaria(
                        profesor_id=pid, dia=dia, franja=franja,
                        tipo='libre', tipo2='media'   # ½ libre solo
                    ))
                    prof_ocupado[pid].add(sl)
                    break

    db.session.commit()
    return len(grupos), grupos_sin_hueco


@app.route('/horarios-construccion/generar-complementarias', methods=['POST'])
@login_required
def generar_complementarias_hc():
    if not current_user.es_admin:
        return redirect(url_for('dashboard'))
    _, sin_hueco = _asignar_grupos_trabajo()
    if sin_hueco:
        flash(f'No se encontró hueco común para: {", ".join(sin_hueco)}. Revisa los horarios de esos profesores.', 'warning')
    else:
        flash('Grupos de trabajo asignados correctamente.', 'success')
    return redirect(url_for('horarios_construccion', tab='horario', vista='profesor'))


@app.route('/horarios-construccion/limpiar-complementarias', methods=['POST'])
@login_required
def limpiar_complementarias_hc():
    if not current_user.es_admin:
        return redirect(url_for('dashboard'))
    SlotComplementaria.query.delete()
    db.session.commit()
    flash('Asignaciones de grupos de trabajo borradas.', 'success')
    return redirect(url_for('horarios_construccion', tab='horario', vista='profesor'))


# ───────────────────────────── REGLAS HORARIO ─────────────────────────────

@app.route('/horarios-construccion/reglas/add', methods=['POST'])
@login_required
def add_regla_hc():
    if not current_user.es_admin:
        return redirect(url_for('dashboard'))
    tipo = request.form.get('tipo', '').strip()
    dureza = request.form.get('dureza', 'dura').strip()
    asig_id = request.form.get('asignatura_id', '') or None
    if asig_id:
        asig_id = int(asig_id)
    prof_id = request.form.get('profesor_id', '') or None
    if prof_id:
        prof_id = int(prof_id)
    curso_id_r = request.form.get('curso_id_regla', '') or None
    if curso_id_r:
        curso_id_r = int(curso_id_r)
    valor = request.form.get('valor', 1, type=int) or 1
    dia = request.form.get('dia', '').strip() or None
    franja = request.form.get('franja', '').strip() or None
    etapa_regla = request.form.get('etapa_regla', '').strip() or None

    if not tipo:
        flash('Tipo de regla requerido.', 'warning')
        return redirect(url_for('horarios_construccion', tab='reglas'))

    TIPOS_PROF = {
        'prof_excluir_curso', 'prof_fijar_curso', 'prof_evitar_curso', 'prof_preferir_curso',
        'prof_excluir_franja', 'prof_fijar_asignatura', 'prof_fijar_curso_asignatura',
        'prof_evitar_franja', 'prof_preferir_asignatura',
        'prof_min_horas', 'prof_max_horas', 'prof_horas_etapa',
    }

    if tipo == 'tutor_clase_etapa':
        if not etapa_regla:
            flash('Selecciona una etapa para esta regla.', 'warning')
            return redirect(url_for('horarios_construccion', tab='reglas'))
        if not asig_id:
            flash('Selecciona una asignatura para esta regla.', 'warning')
            return redirect(url_for('horarios_construccion', tab='reglas'))
    elif tipo in TIPOS_PROF and not prof_id:
        flash('Selecciona un profesor para esta regla.', 'warning')
        return redirect(url_for('horarios_construccion', tab='reglas'))
    elif tipo in ('prof_excluir_curso', 'prof_fijar_curso', 'prof_evitar_curso', 'prof_preferir_curso') and not curso_id_r:
        flash('Selecciona un curso para esta regla.', 'warning')
        return redirect(url_for('horarios_construccion', tab='reglas'))
    elif tipo == 'prof_fijar_curso_asignatura' and (not curso_id_r or not asig_id):
        flash('Selecciona un curso y una asignatura para esta regla.', 'warning')
        return redirect(url_for('horarios_construccion', tab='reglas'))
    elif tipo in ('prof_fijar_asignatura', 'prof_preferir_asignatura') and not asig_id:
        flash('Selecciona una asignatura para esta regla.', 'warning')
        return redirect(url_for('horarios_construccion', tab='reglas'))
    elif tipo in ('prof_excluir_franja', 'prof_evitar_franja') and (not dia or not franja):
        flash('Selecciona día y franja para esta regla.', 'warning')
        return redirect(url_for('horarios_construccion', tab='reglas'))
    elif tipo == 'prof_horas_etapa' and not etapa_regla:
        flash('Selecciona una etapa para esta regla.', 'warning')
        return redirect(url_for('horarios_construccion', tab='reglas'))
    elif tipo in ('max_dia', 'consecutivas') and not asig_id:
        flash('Selecciona una asignatura para esta regla.', 'warning')
        return redirect(url_for('horarios_construccion', tab='reglas'))
    elif tipo == 'fijar_franja' and (not asig_id or not dia or not franja):
        flash('Selecciona asignatura, día y franja para esta regla.', 'warning')
        return redirect(url_for('horarios_construccion', tab='reglas'))

    db.session.add(ReglaHorario(
        tipo=tipo, dureza=dureza,
        asignatura_id=asig_id, profesor_id=prof_id,
        curso_id_regla=curso_id_r,
        valor=valor, dia=dia, franja=franja,
        etapa=etapa_regla
    ))
    db.session.commit()
    flash('Regla añadida.', 'success')
    return redirect(url_for('horarios_construccion', tab='reglas'))


@app.route('/horarios-construccion/reglas/<int:regla_id>/delete', methods=['POST'])
@login_required
def delete_regla_hc(regla_id):
    if not current_user.es_admin:
        return redirect(url_for('dashboard'))
    regla = ReglaHorario.query.get_or_404(regla_id)
    db.session.delete(regla)
    db.session.commit()
    flash('Regla eliminada.', 'success')
    return redirect(url_for('horarios_construccion', tab='reglas'))


# ───────────────────────────── IA HORARIO ─────────────────────────────

@app.route('/horarios-construccion/generar-ia', methods=['POST'])
@login_required
def generar_horario_ia():
    if not current_user.es_admin:
        return redirect(url_for('dashboard'))

    cfg = get_config()
    if not cfg.anthropic_api_key:
        flash('Configura la clave de API de Claude en Configuración antes de usar la IA.', 'danger')
        return redirect(url_for('horarios_construccion', tab='horario'))

    instrucciones = request.form.get('instrucciones_ia', '').strip()

    profesores = Profesor.query.filter_by(activo=True, de_baja=False).all()
    cursos = Curso.query.order_by(Curso.orden).all()
    asignaturas = Asignatura.query.all()
    requisitos = CursoAsignatura.query.all()
    prof_asignaturas = ProfesorAsignatura.query.all()
    franjas_lectivas = [f for f in FRANJAS if f != 'Patio']

    datos_profesores = []
    for p in profesores:
        asigs_ids = {pa.asignatura_id for pa in prof_asignaturas if pa.profesor_id == p.id}
        asigs_nombres = [a.nombre for a in asignaturas if a.id in asigs_ids]
        ocupados = [f"{hp.dia} {hp.franja}" for hp in
                    HorarioProfesor.query.filter_by(profesor_id=p.id, tiene_clase=True).all()]
        datos_profesores.append({
            'id': p.id,
            'nombre': p.nombre,
            'etapa': p.etapa or 'sin especificar',
            'aula_tutoria': p.aula_tutoria or 'ninguna',
            'asignaturas_que_imparte': asigs_nombres,
            'horas_max_semanales': p.horas_max_semanales,
            'franjas_ya_ocupadas': ocupados,
        })

    datos_cursos = []
    for c in cursos:
        if c.aula_cerrada:
            continue
        reqs = [r for r in requisitos if r.curso_id == c.id]
        datos_cursos.append({
            'id': c.id,
            'nombre': c.nombre,
            'etapa': c.etapa or 'sin especificar',
            'asignaturas_requeridas': [
                {'asignatura_id': r.asignatura_id,
                 'nombre': r.asignatura.nombre,
                 'horas_semanales': r.horas_semanales}
                for r in reqs
            ],
        })

    system_prompt = """Eres un asistente experto en creación de horarios escolares.
Tu tarea es generar un horario semanal completo para un colegio.

REGLAS ESTRICTAS:
1. Cada entrada del horario tiene: curso_id, asignatura_id, profesor_id, dia, franja.
2. Los días posibles son: Lunes, Martes, Miércoles, Jueves, Viernes.
3. Las franjas posibles son: 9:00-10:00, 10:00-11:00, 11:30-12:30, 14:30-15:30, 15:30-16:30.
4. Un profesor NO puede estar en dos cursos distintos en la misma franja y día.
5. Un curso solo puede tener UNA asignatura por franja/día.
6. Respeta las horas semanales requeridas de cada asignatura en cada curso.
7. Solo asigna un profesor a una asignatura si ese profesor la puede impartir (está en su lista).
8. Respeta las franjas ya ocupadas de cada profesor (ya tienen clase en esas franjas).
9. Respeta el máximo de horas semanales de cada profesor.
10. Prioriza al tutor del curso para impartir clases en su propio curso cuando sea posible.

RESPONDE ÚNICAMENTE con un array JSON válido (sin texto adicional, sin markdown):
[{"curso_id": 1, "asignatura_id": 2, "profesor_id": 3, "dia": "Lunes", "franja": "9:00-10:00"}, ...]"""

    datos_json = json.dumps({
        'dias': DIAS_SEMANA,
        'franjas': franjas_lectivas,
        'profesores': datos_profesores,
        'cursos': datos_cursos,
    }, ensure_ascii=False, indent=2)

    user_content = f"Datos del colegio:\n{datos_json}"
    if instrucciones:
        user_content += f"\n\nInstrucciones adicionales del administrador:\n{instrucciones}"
    user_content += "\n\nGenera el horario completo en JSON."

    try:
        import anthropic
        client = anthropic.Anthropic(api_key=cfg.anthropic_api_key)
        response = client.messages.create(
            model='claude-sonnet-4-6',
            max_tokens=8192,
            system=[{
                'type': 'text',
                'text': system_prompt,
                'cache_control': {'type': 'ephemeral'},
            }],
            messages=[{'role': 'user', 'content': user_content}],
        )
        raw = response.content[0].text.strip()
        if raw.startswith('```'):
            raw = raw.split('```')[1]
            if raw.startswith('json'):
                raw = raw[4:]
            raw = raw.strip()
        asignaciones = json.loads(raw)
    except Exception as e:
        flash(f'Error al contactar con la IA: {e}', 'danger')
        return redirect(url_for('horarios_construccion', tab='horario'))

    ids_cursos = {c.id for c in cursos}
    ids_asigs = {a.id for a in asignaturas}
    ids_profs = {p.id for p in profesores}

    HorarioAsignacion.query.delete()
    guardados = 0
    vistos = set()
    for item in asignaciones:
        try:
            cid = int(item['curso_id'])
            aid = int(item['asignatura_id'])
            pid = int(item['profesor_id'])
            dia = item['dia']
            franja = item['franja']
        except (KeyError, ValueError):
            continue
        if cid not in ids_cursos or aid not in ids_asigs or pid not in ids_profs:
            continue
        if dia not in DIAS_SEMANA or franja not in franjas_lectivas:
            continue
        key = (cid, dia, franja)
        if key in vistos:
            continue
        vistos.add(key)
        db.session.add(HorarioAsignacion(
            curso_id=cid, asignatura_id=aid, profesor_id=pid, dia=dia, franja=franja
        ))
        guardados += 1

    db.session.commit()
    flash(f'Horario generado por IA: {guardados} franjas asignadas.', 'success')
    return redirect(url_for('horarios_construccion', tab='horario'))


# ───────────────────────────── INIT DB ─────────────────────────────

def init_db():
    db.create_all()
    migrations = [
        'ALTER TABLE profesor ADD COLUMN horas_max_semanales INTEGER DEFAULT 25',
        'ALTER TABLE curso ADD COLUMN etapa VARCHAR(50)',
        'ALTER TABLE curso ADD COLUMN aula_cerrada BOOLEAN DEFAULT 0',
        'ALTER TABLE asignatura ADD COLUMN etapa VARCHAR(50)',
        'ALTER TABLE configuracion_email ADD COLUMN anthropic_api_key VARCHAR(200)',
        'ALTER TABLE profesor ADD COLUMN es_pt BOOLEAN DEFAULT FALSE',
        'ALTER TABLE profesor ADD COLUMN horas_trabajo_personal INTEGER DEFAULT 0',
        "ALTER TABLE profesor ADD COLUMN materias_especiales TEXT DEFAULT '[]'",
        'ALTER TABLE regla_horario ADD COLUMN dureza VARCHAR(10) DEFAULT \'dura\'',
        'ALTER TABLE regla_horario ADD COLUMN profesor_id INTEGER REFERENCES profesor(id)',
        'ALTER TABLE regla_horario ADD COLUMN curso_id_regla INTEGER REFERENCES curso(id)',
        'ALTER TABLE profesor ADD COLUMN horas_libres REAL DEFAULT 0',
        'ALTER TABLE profesor ADD COLUMN horas_lectivas REAL DEFAULT 0',
        'ALTER TABLE profesor ALTER COLUMN horas_libres TYPE REAL USING horas_libres::REAL',
        'ALTER TABLE profesor ALTER COLUMN horas_lectivas TYPE REAL USING horas_lectivas::REAL',
        'ALTER TABLE profesor ADD COLUMN es_educador BOOLEAN DEFAULT FALSE',
        'ALTER TABLE profesor ADD COLUMN horas_pt REAL DEFAULT 0',
        'ALTER TABLE profesor ADD COLUMN horas_educador REAL DEFAULT 0',
        'ALTER TABLE regla_horario ADD COLUMN etapa VARCHAR(50)',
        'ALTER TABLE horario_asignacion ADD COLUMN asignatura2_id INTEGER REFERENCES asignatura(id)',
        'ALTER TABLE horario_asignacion ADD COLUMN profesor2_id INTEGER REFERENCES profesor(id)',
        'ALTER TABLE curso_asignatura ALTER COLUMN horas_semanales TYPE REAL USING horas_semanales::REAL',
        """CREATE TABLE IF NOT EXISTS slot_complementaria (
            id SERIAL PRIMARY KEY,
            profesor_id INTEGER NOT NULL REFERENCES profesor(id),
            dia VARCHAR(20) NOT NULL,
            franja VARCHAR(30) NOT NULL,
            tipo VARCHAR(20) DEFAULT 'libre',
            grupo_id INTEGER REFERENCES grupo_trabajo(id)
        )""",
        # Asignaturas que aparecen en HH y LH deben tener etapa NULL
        "UPDATE asignatura SET etapa = NULL WHERE nombre IN ('Inglés','Religión') AND etapa IS NOT NULL",
        'ALTER TABLE slot_complementaria ADD COLUMN tipo2 VARCHAR(20)',
        'ALTER TABLE horario_asignacion ADD COLUMN es_manual BOOLEAN DEFAULT FALSE',
        'ALTER TABLE slot_complementaria ADD COLUMN es_manual BOOLEAN DEFAULT FALSE',
    ]
    for sql in migrations:
        try:
            with db.engine.connect() as conn:
                conn.execute(db.text(sql))
                conn.commit()
        except Exception:
            pass
    if not ReglaHorario.query.filter_by(tipo='asig_unico_prof').first():
        db.session.add(ReglaHorario(tipo='asig_unico_prof', dureza='dura'))
        db.session.commit()

    # Asegurarse de que el admin no tenga asignaturas ni horario asignado
    admin = Profesor.query.filter_by(es_admin=True).first()
    if admin:
        ProfesorAsignatura.query.filter_by(profesor_id=admin.id).delete()
        HorarioAsignacion.query.filter_by(profesor_id=admin.id).delete()
        db.session.commit()
    else:
        admin = Profesor(
            nombre='Administrador',
            email='admin@colegio.es',
            es_admin=True,
            password_hash=hash_password('admin123')
        )
        db.session.add(admin)
        db.session.commit()
        print('Admin creado: admin@colegio.es / admin123')


# Se ejecuta tanto con gunicorn como con python app.py
with app.app_context():
    init_db()

if __name__ == '__main__':
    app.run(debug=True, port=5000)
